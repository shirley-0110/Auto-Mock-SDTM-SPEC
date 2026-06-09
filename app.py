import streamlit as st
import pandas as pd
import re
import os
import hashlib
import io
import traceback

from io import BytesIO

import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from difflib import get_close_matches
from st_aggrid import AgGrid, GridOptionsBuilder
from datetime import datetime

# Step 2 用到 sas7bdat
try:
    import pyreadstat
    HAS_PYREADSTAT = True
except Exception:
    HAS_PYREADSTAT = False



# ===================================================================================================================================================================================
# 所有 Function
# ===================================================================================================================================================================================

# =================================================================================================================
# 文字處理
# =================================================================================================================
def normalize_text(x):
    if pd.isna(x):
        return ""
    x = str(x) #統一資料型態
    x = x.replace("\n", " ").replace("\r", " ").replace("\xa0", " ") #移除換行
    x = re.sub(r"\s+", " ", x) #壓縮多於空白
    return x.strip().upper()
    # End=========================================================

def normalize_columns(df):
    df = df.copy()
    df.columns = [
        re.sub(r"\s+", " ", str(c).replace("\n", " ").replace("\r", " ").replace("\xa0", " ")).strip()
        for c in df.columns
    ]
    return df
    # End=========================================================


def normalize_date_text(x):
    """
    接受:
      - 2025-09-26
      - 2025/09/26
      - 2025.09.26
      - 20250926
    回傳:
      - 2025-09-26
    """
    if x is None:
        return ""

    s = str(x).strip()
    if not s:
        return ""

    s = s.replace("/", "-").replace(".", "-")

    parts = s.split("-")

    if len(parts) == 3:
        y, m, d = parts
        m = m.zfill(2)
        d = d.zfill(2)
        return f"{y}-{m}-{d}"

    if re.fullmatch(r"\d{8}", s):
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"

    return s
    # End=========================================================


# =================================================================================================================
# 匯入/匯出Excel各種工具
# =================================================================================================================
def find_column(columns, required_keywords):
    for col in columns:
        upper_col = normalize_text(col)
        if all(k.upper() in upper_col for k in required_keywords):
            return col
    return None
    # End=========================================================


def row_contains_keywords(row_values, keyword_groups):
    cells = [normalize_text(v) for v in row_values]

    for cell in cells:
        for group in keyword_groups:
            if all(k.upper() in cell for k in group):
                return True
    return False
    # End=========================================================


def detect_header_row(file_bytes, sheet_name, keyword_groups, max_scan_rows=30):
    raw_df = pd.read_excel(
        BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=None,
        nrows=max_scan_rows,
        dtype=str
    )

    for idx, row in raw_df.iterrows():
        if row_contains_keywords(row.tolist(), keyword_groups):
            return idx

    return None
    # End=========================================================


def read_sheet_with_detected_header(
    file_bytes,
    sheet_name,
    keyword_groups,
    max_scan_rows=30
):
    header_row_zero_based = detect_header_row(
        file_bytes=file_bytes,
        sheet_name=sheet_name,
        keyword_groups=keyword_groups,
        max_scan_rows=max_scan_rows
    )

    if header_row_zero_based is None:
        raise ValueError(f"無法自動判斷 {sheet_name} 的 header row")

    df = pd.read_excel(
        BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=header_row_zero_based
    )
    df = normalize_columns(df)

    return df, header_row_zero_based + 1
    # End=========================================================



def Export_excel(sheet_dict):
    output = BytesIO()

    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheet_dict.items():

            # 如果是空 df 就 skip（避免空 sheet）
            if df is None or df.empty:
                continue

            df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.book[sheet_name]

            # Freeze first row
            ws.freeze_panes = "A2"

            # Auto filter（非常推薦）
            ws.auto_filter.ref = ws.dimensions

            # Header style
            header_fill = PatternFill(
                start_color="F4A300",
                end_color="F4A300",
                fill_type="solid"
            )

            header_font = Font(name="Calibri", size=10, bold=True)
            normal_font = Font(name="Calibri", size=10)

            align_wrap_top = Alignment(wrap_text=True, vertical="top")
            align_wrap_center = Alignment(wrap_text=True, vertical="center")

            # Header style
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_wrap_center

            # Body style（避免重複設定 header）
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = normal_font
                    cell.alignment = align_wrap_top

            # 欄寬自動（稍微 smarter）
            for col_idx, col in enumerate(ws.columns, start=1):
                max_length = 0

                col_letter = get_column_letter(col_idx)

                for cell in col:
                    try:
                        if cell.value:
                            val = str(cell.value)
                            max_length = max(max_length, len(val))
                    except:
                        pass

                # 限制最大寬度（避免爆炸）
                ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

            # Row height（改善 wrap）
            for row in ws.iter_rows():
                max_lines = 1

                for cell in row:
                    if cell.value:
                        text = str(cell.value)
                        approx_lines = max(
                            text.count("\n") + 1,
                            (len(text) // 40) + 1
                        )
                        max_lines = max(max_lines, approx_lines)

                ws.row_dimensions[row[0].row].height = max_lines * 15

    output.seek(0)
    return output.getvalue()
    # End=========================================================



# 抓檔名 (Sponsor, Protocol)
def extract_protocol_no_from_filename(file_name):
    if not file_name:
        return ""

    name = os.path.splitext(file_name)[0].strip()

    parts = re.split(r"ecrf\s*schema", name, flags=re.IGNORECASE)
    prefix = parts[0].strip().strip("_") if parts else name

    tokens = [t.strip() for t in prefix.split("_") if t.strip()]

    sponsor = ""
    protocol = ""

    if len(tokens) >= 2:
        sponsor = tokens[0]
        protocol = tokens[1]
    elif len(tokens) == 1:
        protocol = tokens[0]

    return sponsor, protocol
    # End=========================================================




# 處理OID
def extract_form_oids(series):
    domains = set()

    for value in series.dropna():
        text = str(value).strip()
        if not text:
            continue

        parts = re.split(r"[,\n;/]+", text)

        for part in parts:
            item = part.strip()
            if item:
                domains.add(item.upper())

    return domains
    # End=========================================================



# 整個eCRF schema匯入+暫存
def build_step1_context(file_bytes, all_sheets):

    # 1. 匯入SoA
    soa_df, _ = read_sheet_with_detected_header(
        file_bytes=file_bytes,
        sheet_name="SoA",
        keyword_groups=[
            ["FORM", "OID"],
            ["ABBREVIATION"]
        ]
    )

    form_oid_col = find_column(soa_df.columns, ["FORM", "OID"])
    
    # fallback：如果沒有 FORM OID，用 Abbreviation
    if form_oid_col is None:
        form_oid_col = find_column(soa_df.columns, ["ABBREVIATION"])

    if form_oid_col is None:
        raise ValueError("SoA 分頁中找不到 Form OID 或 Abbreviation 欄位")

    valid_domains = extract_form_oids(soa_df[form_oid_col])

    sheet_upper_map = {s.upper(): s for s in all_sheets}

    available_sheets = [
        sheet_upper_map[d] for d in valid_domains if d in sheet_upper_map
    ]

    missing_sheets = [
        d for d in valid_domains if d not in sheet_upper_map
    ]


    # 2. Folder
    folder_df, _ = read_sheet_with_detected_header(
        file_bytes=file_bytes,
        sheet_name="Folder",
        keyword_groups=[["ABBREVIATION"], ["FULL", "TERM"]]
    )


    # 3. CRF Domain Sheets
    domain_df_map = {}
    sheet_errors = []

    for sheet in available_sheets:

        try:
            df, _ = read_sheet_with_detected_header(
                file_bytes=file_bytes,
                sheet_name=sheet,
                keyword_groups=[["SDTM", "TARGET"]]
            )
            domain_df_map[sheet] = df

        except Exception:
            sheet_errors.append(sheet)

    return {
        "soa_df": soa_df,
        "folder_df": folder_df,
        "domain_df_map": domain_df_map,
        "available_sheets": available_sheets,
        "missing_sheets": missing_sheets,
        "sheet_errors": sheet_errors
    }
    # End=========================================================



# =================================================================================================================
# 匯入/整理Config
# =================================================================================================================
def load_domains_config(version):
    if version == "Version 3.3":
        path = "config/v33/domains.sas7bdat"
    else:
        path = "config/v34/domains.sas7bdat"

    if not HAS_PYREADSTAT:
        raise ImportError("目前環境尚未安裝 pyreadstat，請先在 requirements.txt 加入 pyreadstat")

    if not os.path.exists(path):
        raise FileNotFoundError(f"找不到 config 檔：{path}")

    cfg_df, _ = pyreadstat.read_sas7bdat(path)
    cfg_df = normalize_columns(cfg_df)

    return cfg_df, path
    # End=========================================================


def standardize_domains_config(cfg_df):
    df = cfg_df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]

    rename_map = {
        "domain": "Dataset",
        "dlabel": "Dataset Label",
        "repeat": "Repeat",
        "refdata": "RefData",
        "structure": "Structure",
        "keyvars": "Key Variables",
        "keyseq": "KeySeq",
        "varnum": "VarNum",
        "name": "Variable",
        "label": "Variable Label",
        "type": "Data Type",
        "mandatory": "Mandatory",
        "role": "Role",
        "core": "Core",
        "ctcode": "CT Code",
        "class": "Class"
    }

    df = df.rename(columns=rename_map)

    if "Dataset" in df.columns:
        df["Dataset"] = df["Dataset"].astype(str).str.upper().str.strip()

    if "Variable" in df.columns:
        df["Variable"] = df["Variable"].astype(str).str.upper().str.strip()

    
    # 特殊處理：STRTPT / ENRTPT → STENRF
    if "CT Code" in df.columns:
        df["CT Code"] = df["CT Code"].replace({
            "STRTPT": "STENRF",
            "ENRTPT": "STENRF"
        })

    return df
    # End=========================================================


# =================================================================================================================
# 特定使用
# =================================================================================================================

# 抓SoA的Visit
def build_soa_visit_list(soa_df, folder_df):
    """
    從 SoA + Folder 建立 SoA List:
      CRF Dataset / Abbreviation / Visit

    規則：
      - SoA 的 row = Source CRF Sheet (Form OID)
      - SoA 的 visit 欄位只要 cell = X，就輸出一列
      - Folder 的 Abbreviation -> Full Term 對出 Visit
    """

    # 1 呼叫SoA
    form_oid_col = find_column(soa_df.columns, ["FORM", "OID"])
    if form_oid_col is None:
        form_oid_col = find_column(soa_df.columns, ["ABBREVIATION"])

    if form_oid_col is None:
        raise ValueError("SoA 分頁中找不到 Form OID 或 Abbreviation 欄位")


    # SoA 所有欄位
    soa_columns = [str(c).strip() for c in soa_df.columns if str(c).strip()]

    # 這些欄位不是 visit abbreviation
    non_visit_headers = {
        "FORM OID", "FORMOID", "FORM",
        "CRF NAME", "FORM NAME",
        "DESCRIPTION", "SEQ", "ORDER"
    }

    visit_cols = [
        (col, idx)
        for idx, col in enumerate(soa_columns)
        if normalize_text(col) not in non_visit_headers
    ]
    
    # 2 呼叫 Folder
    abbr_col = find_column(folder_df.columns, ["ABBREVIATION"])
    if abbr_col is None:
        raise ValueError("Folder 分頁中找不到 Abbreviation 欄位")

    full_term_col = find_column(folder_df.columns, ["FULL", "TERM"])
    if full_term_col is None:
        raise ValueError("Folder 分頁中找不到 Full Term 欄位")

    folder_work = (
        folder_df[[abbr_col, full_term_col]]
        .rename(columns={abbr_col: "Abbreviation", full_term_col: "Visit"})
        .dropna()
    )
    
    folder_work["Abbreviation"] = folder_work["Abbreviation"].astype(str).str.strip().str.upper()
    folder_work["Visit"] = folder_work["Visit"].astype(str).str.strip()

    folder_lookup = dict(
        folder_work.drop_duplicates("Abbreviation")[["Abbreviation", "Visit"]].values
    )

    # -----------------------------
    # 3) 展開 SoA List
    # -----------------------------
    records = []

    for _, row in soa_df.iterrows():
        source_sheet = str(row.get(form_oid_col, "")).strip().upper()

        if not source_sheet:
            continue

        for col, col_idx in visit_cols:
            abbr = normalize_text(col)
            abbr = re.sub(r'[\*\^]+', '', abbr).strip()
            cell_val = str(row.get(col, "")).strip().upper()

            # 只抓 ticked X
            if str(cell_val).strip().upper() in ["X", "TRUE", "1"]:
                visit_name = folder_lookup.get(abbr, "")

                records.append({
                    "CRF Dataset": source_sheet,
                    "Abbreviation": abbr,
                    "Visit": visit_name,
                    "Visit_order": col_idx
                })

    if records:
        out_df = pd.DataFrame(records).drop_duplicates()
        
        # 排序：先 CRF Dataset，再 Visit_order
        out_df = out_df.sort_values(
            by=["CRF Dataset", "Visit_order"],
            ascending=[True, True]
        ).reset_index(drop=True)

    else:
        out_df = pd.DataFrame(columns=[
            "CRF Dataset", "Abbreviation", "Visit", "Visit_order"
        ])

    return out_df
    # End=========================================================



# 抓各個CRF Domain的Field OID
def find_source_variable_column(columns):
    """
    Source CRF Variable 優先抓 Field OID
    """
    priority_exact = [
        "FIELD OID",
        "FIELDOID",
        "FIELD OID NAME",
        "CRF FIELD OID",
        "SOURCE FIELD OID",
        "VARIABLE",
        "VARIABLE NAME",
        "CRF VARIABLE",
        "SOURCE VARIABLE"
    ]

    normalized_map = {col: normalize_text(col) for col in columns}

    for target in priority_exact:
        for col, norm_col in normalized_map.items():
            if norm_col == target:
                return col

    for col, norm_col in normalized_map.items():
        if "FIELD" in norm_col and "OID" in norm_col:
            return col

    for col, norm_col in normalized_map.items():
        if "VARIABLE" in norm_col and "TARGET" not in norm_col and "SDTM" not in norm_col:
            return col

    return None
    # End=========================================================



# 處理SDTM IG Target
def parse_sdtm_targets(value):
    """
    規則：
      - 只用分號 ; 和換行切
      - 不用逗號和斜線切
      - 支援：
          AE.AETERM
          VS.VSTESTCD="TEMP"
          DM.SEX='F'
    """
    parsed_records = []
    unparsed_tokens = []

    if pd.isna(value):
        return parsed_records, unparsed_tokens

    text = str(value).strip()
    if not text:
        return parsed_records, unparsed_tokens

    tokens = re.split(r"[;\n]+", text)

    pattern = re.compile(
        r'^\s*([A-Za-z][A-Za-z0-9]{0,7})\s*\.\s*([A-Za-z_][A-Za-z0-9_]*)\s*(?:=\s*["\']?(.*?)["\']?)?\s*$'
    )

    for token in tokens:
        token = token.strip()
        if not token:
            continue

        match = pattern.match(token)
        if match:
            dom, var, assign_val = match.groups()

            if assign_val is None:
                assign_val = ""
            else:
                assign_val = str(assign_val).strip()

            parsed_records.append({
                "SDTM Domain": dom.upper(),
                "SDTM Variable": var.upper(),
                "Assign Value": assign_val
            })
        else:
            unparsed_tokens.append(token)

    return parsed_records, unparsed_tokens
    # End=========================================================



# 處理CRF -> SDTM Variable Mapping
def build_sdtm_mapping(domain_df_map):

    mapping_records = []
    detail_records = []
    sheet_errors = []
    unparsed_records = []

    for sheet, df in domain_df_map.items():
        target_col = find_column(df.columns, ["SDTM", "TARGET"])
        source_var_col = find_source_variable_column(df.columns)
        source_dtype_col = find_column(df.columns, ["FIELD", "TYPE"])

        if target_col is None:
            sheet_errors.append(sheet)
            continue

        for _, row in df.iterrows():
            raw_target = row[target_col]
            source_var = row[source_var_col] if source_var_col is not None else ""
            source_dtype = row[source_dtype_col] if source_dtype_col is not None else ""
        
            parsed_records, unparsed_tokens = parse_sdtm_targets(raw_target)

            for rec in parsed_records:
                mapping_records.append({
                    "SDTM Domain": rec["SDTM Domain"],
                    "SDTM Variable": rec["SDTM Variable"]
                })

                detail_records.append({
                    "CRF Dataset": sheet,
                    "CRF Variable": source_var,
                    "CRF Data Type": source_dtype,
                    "SDTM Domain": rec["SDTM Domain"],
                    "SDTM Variable": rec["SDTM Variable"],
                    "Assign Value": rec["Assign Value"],
                    "SDTM IG Target Raw": raw_target
                })

            for token in unparsed_tokens:
                if str(token).strip():
                    unparsed_records.append({
                        "CRF Dataset": sheet,
                        "CRF Variable": source_var,
                        "CRF Data Type": source_dtype,
                        "SDTM IG Target Raw": raw_target,
                        "Unparsed Token": token
                    })

    mapping_df = (
        pd.DataFrame(mapping_records)
        .drop_duplicates()
        .sort_values(by=["SDTM Domain", "SDTM Variable"])
        .reset_index(drop=True)
    ) if mapping_records else pd.DataFrame(columns=["SDTM Domain", "SDTM Variable"])

    detail_df = (
        pd.DataFrame(detail_records)
        .drop_duplicates()
        .sort_values(["SDTM Domain", "SDTM Variable", "CRF Dataset"])
        .reset_index(drop=True)
    ) if detail_records else pd.DataFrame()

    return mapping_df, detail_df, sheet_errors, unparsed_records
    # End=========================================================




# 整批CRF處理
def process_uploaded_excel(file_bytes, all_sheets):

    # Step1 context
    ctx = build_step1_context(file_bytes, all_sheets)

    soa_df = ctx["soa_df"]
    folder_df = ctx["folder_df"]
    domain_df_map = ctx["domain_df_map"]


    # SoA list
    soa_list_df = build_soa_visit_list(soa_df, folder_df)

    # SDTM mapping
    mapping_df, detail_df, mapping_errors, unparsed_records = build_sdtm_mapping(
        domain_df_map
    )
    
    # CT mapping
    ct_mapping_df, ct_mapping_sheet_errors = build_ct_mapping_seed(
        domain_df_map,
        st.session_state["var_to_ctcode"]
    )

    return {
        "soa_list_df": soa_list_df,
        "mapping_df": mapping_df,
        "detail_df": detail_df,
        "unparsed_records": unparsed_records,
        "mapping_sheet_errors": mapping_errors,
        "ct_mapping_df": ct_mapping_df,
        "ct_mapping_sheet_errors": ct_mapping_sheet_errors,
        "available_sheets": ctx["available_sheets"],
        "missing_sheets": ctx["missing_sheets"],
        "sheet_errors": ctx["sheet_errors"]
    }
    # End=========================================================





# =================================================================================================================
# CT系列
# =================================================================================================================

# 抓CRF Option Displayed Value欄位
def find_option_displayed_value_column(columns):
    """
    優先抓 Option Displayed Value
    """
    priority_exact = [
        "OPTION DISPLAYED VALUE",
        "OPTION_DISPLAYED_VALUE",
        "OPTION DISPLAY VALUE",
        "DISPLAYED VALUE",
        "OPTION VALUE",
        "OPTIONS"
    ]

    normalized_map = {col: normalize_text(col) for col in columns}

    # 1. 先找完全一致
    for target in priority_exact:
        for col, norm_col in normalized_map.items():
            if norm_col == target:
                return col

    # 2. 再找包含 OPTION + DISPLAY + VALUE
    for col, norm_col in normalized_map.items():
        if "OPTION" in norm_col and "DISPLAY" in norm_col and "VALUE" in norm_col:
            return col

    # 3. 再退一步找 OPTION + VALUE
    for col, norm_col in normalized_map.items():
        if "OPTION" in norm_col and "VALUE" in norm_col:
            return col

    return None
    # End=========================================================


# 處理CRF Option Displayed Value
def split_option_displayed_values(value):
    """
    將 Option Displayed Value 拆成多個 option
    規則：
      - 只用分號 ; 和換行切
      - 不用逗號和斜線切
    """
    if pd.isna(value):
        return []

    text = str(value).strip()
    if not text:
        return []

    tokens = re.split(r"[;\n]+", text)

    out = []
    for token in tokens:
        token = str(token).strip()
        if token:
            out.append(token)

    return out
    # End=========================================================



# 處理CRF -> SDTM CT Mapping
def build_ct_mapping_seed(domain_df_map, var_to_ctcode):
    """
    從已讀入的 CRF Domain DataFrames 建立 CT Mapping Seed

    輸出：
      - ct_mapping_df
      - ct_mapping_sheet_errors
    """

    seed_records = []
    ct_mapping_sheet_errors = []

    for sheet, df in domain_df_map.items():

        try:
            target_col = find_column(df.columns, ["SDTM", "TARGET"])
            source_var_col = find_source_variable_column(df.columns)
            option_col = find_option_displayed_value_column(df.columns)

            # 至少要有 SDTM Target 與 Source Variable
            if target_col is None or source_var_col is None:
                ct_mapping_sheet_errors.append(sheet)
                continue

            for _, row in df.iterrows():
                try:
                    raw_target = row.get(target_col, "")
                    source_var = row.get(source_var_col, "")
                    raw_option = row.get(option_col, "")

                    source_var = "" if pd.isna(source_var) else str(source_var).strip()
                    raw_target = "" if pd.isna(raw_target) else str(raw_target).strip()
                    raw_option = "" if pd.isna(raw_option) else str(raw_option).strip()

                    # 沒有 source var 就跳過
                    if not source_var:
                        continue

                    # 先 parse SDTM target
                    parsed_records, _ = parse_sdtm_targets(raw_target)

                    # 沒 parse 到 SDTM target，就不進 seed
                    if not parsed_records:
                        continue

                    # 拆 options
                    option_tokens = split_option_displayed_values(raw_option)

                    for rec in parsed_records:

                        sdtm_domain = str(rec.get("SDTM Domain", "")).strip().upper()
                        sdtm_var = str(rec.get("SDTM Variable", "")).strip().upper()

                        ctcode = var_to_ctcode.get(sdtm_var, "")
                        ctcode = "" if pd.isna(ctcode) else str(ctcode).strip().upper()

                        assign_val = rec.get("Assign Value", "")
                        assign_val = "" if pd.isna(assign_val) else str(assign_val).strip()

                        # -------------------------------------------------
                        # 情況 A：有 Assign Value
                        #   → 優先保留 assign
                        # -------------------------------------------------
                        if assign_val:

                            seed_records.append({
                                "SDTM Domain": sdtm_domain,
                                "SDTM Variable": sdtm_var,
                                "CT Code": ctcode,
                                "Assign Value": assign_val,
                                "CRF Option Value": "",
                                "Original Value": assign_val,
                                "Original Value Normalized": normalize_text(assign_val)
                            })

                        # -------------------------------------------------
                        # 情況 B：沒有 Assign Value，但有 option
                        #   → 逐個 option 展開
                        # -------------------------------------------------
                        elif option_tokens:

                            for opt in option_tokens:
                                opt = "" if pd.isna(opt) else str(opt).strip()
                                if not opt:
                                    continue

                                seed_records.append({
                                    "SDTM Domain": sdtm_domain,
                                    "SDTM Variable": sdtm_var,
                                    "CT Code": ctcode,
                                    "Assign Value": "",
                                    "CRF Option Value": opt,
                                    "Original Value": opt,
                                    "Original Value Normalized": normalize_text(opt)
                                })

                        # -------------------------------------------------
                        # 情況 C：既沒有 Assign，也沒有 option
                        #   → 如果有 CT Code，仍保留一列供後續 CT 展開
                        # -------------------------------------------------
                        elif ctcode:

                            seed_records.append({
                                "SDTM Domain": sdtm_domain,
                                "SDTM Variable": sdtm_var,
                                "CT Code": ctcode,
                                "Assign Value": "",
                                "CRF Option Value": "",
                                "Original Value": "",
                                "Original Value Normalized": ""
                            })

                except Exception:
                    # 單列失敗不影響整張 sheet
                    continue

        except Exception:
            ct_mapping_sheet_errors.append(sheet)
            continue

    if seed_records:
        ct_mapping_df = (
            pd.DataFrame(seed_records)
            .drop_duplicates(
                subset=[
                    "SDTM Domain",
                    "SDTM Variable",
                    "CT Code",
                    "Assign Value",
                    "CRF Option Value",
                    "Original Value",
                    "Original Value Normalized"
                ]
            )
            .sort_values(
                by=[
                    "SDTM Domain",
                    "SDTM Variable",
                    "CT Code",
                    "Assign Value",
                    "CRF Option Value",
                    "Original Value"
                ]
            )
            .reset_index(drop=True)
        )
    else:
        ct_mapping_df = pd.DataFrame(columns=[
            "SDTM Domain",
            "SDTM Variable",
            "CT Code",
            "Assign Value",
            "CRF Option Value",
            "Original Value",
            "Original Value Normalized"
        ])

    return ct_mapping_df, sorted(list(set(ct_mapping_sheet_errors)))

    # End=========================================================





def standardize_ct_mapping_dict(df):
    df = df.copy()
    df = normalize_columns(df)

    rename_map = {
        "ID": "CT Code",
        "ORIVAL": "Original Value Normalized",
        "CTVAL": "CT Term",
        "ACTIVE": "Active",
        "NOTE": "Notes"
    }

    # 先把欄名 normalize 到大寫，再對映
    normalized_col_map = {c: normalize_text(c) for c in df.columns}
    new_cols = {}
    for original_col, norm_col in normalized_col_map.items():
        if norm_col in rename_map:
            new_cols[original_col] = rename_map[norm_col]

    df = df.rename(columns=new_cols)

    required_cols = ["CT Code", "Original Value Normalized", "CT Term"]
    for c in required_cols:
        if c not in df.columns:
            raise ValueError(f"CT Mapping Dictionary 缺少必要欄位: {c}")

    df["CT Code"] = df["CT Code"].astype(str).str.strip().str.upper()
    df["Original Value Normalized"] = df["Original Value Normalized"].astype(str).str.strip().str.upper()
    df["CT Term"] = df["CT Term"].astype(str).str.strip()

    if "Active" in df.columns:
        df["Active"] = df["Active"].astype(str).str.strip().str.upper()
        df = df[df["Active"] != "N"]

    df = df[
        (df["CT Code"] != "") &
        (df["Original Value Normalized"] != "") &
        (df["CT Term"] != "")
    ].drop_duplicates(subset=["CT Code", "Original Value Normalized"], keep="first")

    return df
    # End=========================================================




def build_ct_mapping(ct_seed_df, mapping_dict_df, ct_alias_df=None):
    """
    依 CT Code + Original Variable Normalized 做 CT mapping

    Parameters
    ----------
    ct_seed_df : DataFrame
        來自 build_ct_mapping_seed() 的輸出
        必要欄位:
        - SDTM Domain
        - SDTM Variable
        - CT Code
        - Original Value
        - Original Value Normalized

    mapping_dict_df : DataFrame
        開發者維護的 CT Mapping Dictionary
        必要欄位:
        - CT Code
        - Original Value Normalized
        - CT Term

    ct_alias_df : DataFrame, optional
        CT 主檔整理出的 alias 表，用來 fallback
        建議欄位:
        - CT Code
        - Original Value Normalized
        - CT Term
        - Alias Source   (optional)

    Returns
    -------
    matched_df : DataFrame
    unmatched_df : DataFrame
    """

    if ct_seed_df is None or ct_seed_df.empty:
        empty_cols = [
            "SDTM Domain", "SDTM Variable", "CT Code", "Original Value", "Original Value Normalized",
            "CT Term", "Match Method"
        ]
        return pd.DataFrame(columns=empty_cols), pd.DataFrame(columns=empty_cols)


    # 統一 seed 欄名（如果 normalize_columns 後有大小寫差異）
    seed = ct_seed_df.copy()
    seed = normalize_columns(seed)
    
    seed_col_map = {c: normalize_text(c) for c in seed.columns}
    rename_seed = {}
    for original, norm in seed_col_map.items():
        if norm == "SDTM DOMAIN":
            rename_seed[original] = "SDTM Domain"
        elif norm == "SDTM VARIABLE":
            rename_seed[original] = "SDTM Variable"
        elif norm == "CTCODE":
            rename_seed[original] = "CT Code"
        elif norm == "ORIGINAL VALUE":
            rename_seed[original] = "Original Value"
        elif norm == "ORIGINAL VALUE NORMALIZED":
            rename_seed[original] = "Original Value Normalized"

    seed = seed.rename(columns=rename_seed)

    required_seed_cols = [
        "SDTM Domain", "SDTM Variable", "CT Code", "Original Value", "Original Value Normalized"
    ]
    for c in required_seed_cols:
        if c not in seed.columns:
            raise ValueError(f"CT Seed 缺少必要欄位: {c}")

    seed["CT Code"] = seed["CT Code"].astype(str).str.strip().str.upper()
    seed["Original Value Normalized"] = seed["Original Value Normalized"].astype(str).str.strip().str.upper()


    # -------------------------------------------------
    # Step 1：分流（核心設計）
    # -------------------------------------------------
    no_ct_mask = seed["CT Code"] == ""

    # DERIVED（無 CTcode）
    derived_df = seed[no_ct_mask].copy()
    derived_df["CT Term"] = derived_df["Original Value"]
    derived_df["Match Method"] = "DERIVED"

    # 有 CTcode 才做 mapping
    seed_ct = seed[~no_ct_mask].copy()

    # -------------------------------------------------
    # Step 2：DICT mapping
    # -------------------------------------------------
    mapping_dict = standardize_ct_mapping_dict(mapping_dict_df)

    mapping_dict["CT Code"] = mapping_dict["CT Code"].astype(str).str.strip().str.upper()
    mapping_dict["Original Value Normalized"] = mapping_dict["Original Value Normalized"].apply(normalize_text)
    mapping_dict["CT Term"] = mapping_dict["CT Term"].astype(str).str.strip()

    mapped = seed_ct.merge(
        mapping_dict[["CT Code", "Original Value Normalized", "CT Term"]],
        how="left",
        on=["CT Code", "Original Value Normalized"]
    )

    mapped["Match Method"] = mapped["CT Term"].apply(
        lambda x: "DICT" if pd.notna(x) and str(x).strip() != "" else "UNMATCHED"
    )

    # -------------------------------------------------
    # Step 3：合併結果
    # -------------------------------------------------
    final_df = pd.concat([mapped, derived_df], ignore_index=True)

    # -------------------------------------------------
    # Step 4：split matched / unmatched
    # -------------------------------------------------
    matched_df = final_df[final_df["Match Method"] != "UNMATCHED"].copy()
    unmatched_df = final_df[final_df["Match Method"] == "UNMATCHED"].copy()

    # -------------------------------------------------
    # Step 5：排序
    # -------------------------------------------------
    sort_cols = [
        "SDTM Domain",
        "SDTM Variable",
        "CT Code",
        "Original Value Normalized"
    ]

    matched_df = matched_df.sort_values(sort_cols).reset_index(drop=True)
    unmatched_df = unmatched_df.sort_values(sort_cols).reset_index(drop=True)

    return matched_df, unmatched_df

    # End=========================================================



# =================================================================================================================
# 根據所選的CT版本 load SDTM Terminology
# =================================================================================================================
def load_ct_master(sdtm_ct):

    base_path = "config/sdtm_ct"
    
    file_name = f"SDTM Terminology {sdtm_ct}.txt"
    file_path = os.path.join(base_path, file_name)

    # 讀 txt
    df = pd.read_csv(file_path, sep="\t", dtype=str).fillna("")
    df = normalize_columns(df)
    
    rename_map = {}

    for col in df.columns:
        ncol = normalize_text(col)

        if ncol in ["CODELIST CODE", "CODE LIST CODE", "NCI CODELIST CODE"]:
            rename_map[col] = "Codelist Code"

        elif ncol in ["CODELIST NAME"]:
            rename_map[col] = "Codelist Name"

        elif ncol in ["CDISC SUBMISSION VALUE", "SUBMISSION VALUE"]:
            rename_map[col] = "Submission Value"

        elif ncol in ["NCI CODE", "NCI TERM CODE", "CODE"]:
            rename_map[col] = "Code"

    df = df.rename(columns=rename_map)
    df = df.loc[:, ~df.columns.duplicated()]

    
    # 保底欄位
    for c in ["Codelist Code", "Codelist Name", "Submission Value", "Code"]:
        if c not in df.columns:
            df[c] = ""

    df["Codelist Name"] = (
        df["Codelist Name"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df["Submission Value"] = (
        df["Submission Value"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df["ID_Temp"] = df["Codelist Name"]

    return df.reset_index(drop=True), {
        "source_type": "local",
        "resolved_version": sdtm_ct,
        "file_path": file_path,
        "status": "success"
    }
    # End=========================================================


# =================================================================================================================
# Step 2 - Mock SDTM SPEC
# =================================================================================================================
def build_trial_design_sheets(protocol_no, protocol_title, sdtm_version, sdtm_ct, snomed_version, medrt_version, unii_version, unique_visit_df, ct_master_df=None):

    # ----------------------------------------
    # Normalize inputs
    # ----------------------------------------
    protocol_no = str(protocol_no or "").strip()
    protocol_title = str(protocol_title or "").strip()
    std_ver = str(sdtm_version).upper().replace("VERSION", "").strip()

    sdtmver_map = {
        "3.4": "2.0",
        "3.3": "1.7"
    }

    # ----------------------------------------
    # TS TSPARMCD list
    # ----------------------------------------
    tsparmcd_list = [
        "ACTSUB","ADAPT","ADDON","AGEMAX","AGEMIN","DCUTDESC","DCUTDTC","EXTTIND",
        "FCNTRY","HLTSUBJI","INDIC","INTMODEL","INTTYPE","LENGTH","NARMS","NCOHORT",
        "OBJPRIM","OBJSEC","ONGOSIND","OUTMSPRI","PCLAS","PDPSTIND","PDSTIND",
        "PIPIND","PLANSUB","RANDOM","RDIND","REGID","SDTIGVER","SDTMVER","SENDTC",
        "SEXPOP","SPONSOR","SSTDTC","STOPRULE","STYPE","TBLIND","TCNTRL","TDIGRP",
        "THERAREA","TINDTP","TITLE","TPHASE","TRT","TTYPE"
    ]

    tsval_map = {
        "SDTIGVER": std_ver,
        "SDTMVER": sdtmver_map.get(std_ver, ""),
        "TITLE": protocol_title
    }


    refver_map = {
        "ADAPT":   ("CDISC CT", sdtm_ct),
        "ADDON":   ("CDISC CT", sdtm_ct),
        "EXTTIND": ("CDISC CT", sdtm_ct),
        "HLTSUBJI": ("CDISC CT", sdtm_ct),
        "INTMODEL": ("CDISC CT", sdtm_ct),
        "INTTYPE": ("CDISC CT", sdtm_ct),
        "ONGOSIND": ("CDISC CT", sdtm_ct),
        "PDPSTIND": ("CDISC CT", sdtm_ct),
        "PDSTIND": ("CDISC CT", sdtm_ct),
        "PIPIND": ("CDISC CT", sdtm_ct),        
        "RANDOM": ("CDISC CT", sdtm_ct),
        "RDIND": ("CDISC CT", sdtm_ct),
        "SEXPOP": ("CDISC CT", sdtm_ct),
        "STYPE": ("CDISC CT", sdtm_ct),
        "TBLIND": ("CDISC CT", sdtm_ct),
        "TCNTRL": ("CDISC CT", sdtm_ct),
        "TINDTP": ("CDISC CT", sdtm_ct),
        "TPHASE": ("CDISC CT", sdtm_ct),
        "TTYPE": ("CDISC CT", sdtm_ct),

        "INDIC":   ("SNOMED", snomed_version),
        "TDIGRP":  ("SNOMED", snomed_version),

        "PCLAS":   ("MED-RT", medrt_version),

        "TRT":     ("UNII", unii_version),

        "SPONSOR": ("D-U-N-S NUMBER", ""),
        "REGID":   ("ClinicalTrials.gov", ""),
        "FCNTRY":  ("ISO 3166", "")
    }

    # ----------------------------------------
    # 1. TA / TE 建立空架構）
    # ----------------------------------------
    ta_df = pd.DataFrame([{
        "STUDYID": protocol_no,
        "DOMAIN": "TA",
        "ARMCD": "",
        "ARM": "",
        "TAETORD": "",
        "ETCD": "",
        "ELEMENT": "",
        "TABRANCH": "",
        "TATRANS": "",
        "EPOCH": ""
    }])

    te_df = pd.DataFrame([{
        "STUDYID": protocol_no,
        "DOMAIN": "TE",
        "ETCD": "",
        "ELEMENT": "",
        "TESTRL": "",
        "TEENRL": "",
        "TEDUR": ""
    }])


    ti_df = pd.DataFrame(columns=[
        "STUDYID","DOMAIN","IETESTCD","IETEST",
        "IECAT","TIVERS"
    ])

    # ----------------------------------------
    # TI
    # ----------------------------------------
    ti_rows = []

    categories = ["INCLUSION", "EXCLUSION"]
    for i, cat in enumerate(categories, start=1):
        
        ti_rows.append({
            "STUDYID": protocol_no,
            "DOMAIN": "TI",
            "IETESTCD": "",
            "IETEST": "",
            "IECAT": cat,
            "TIVERS": ""
        })
    ti_df = pd.DataFrame(ti_rows)


    # ----------------------------------------
    # TS TSPARMCD -> TSPARM lookup（從 CT master 抓）
    # ----------------------------------------
    tsparm_lookup = {}

    if ct_master_df is not None and not ct_master_df.empty:

        ct = ct_master_df.copy()

        # 保底欄位
        for col in ["Code", "Codelist Name", "Submission Value"]:
            if col not in ct.columns:
                ct[col] = ""

        ct["Code"] = (
            ct["Code"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        ct["Codelist Name"] = (
            ct["Codelist Name"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        ct["Submission Value"] = (
            ct["Submission Value"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        # Code side: Trial Summary Parameter Test Code
        tsparmcd_df = ct[
            ct["Codelist Name"] == "Trial Summary Parameter Test Code"
        ][["Code", "Submission Value"]].copy()

        tsparmcd_df = tsparmcd_df.rename(columns={
            "Submission Value": "TSPARMCD"
        })

        tsparmcd_df["TSPARMCD"] = tsparmcd_df["TSPARMCD"].str.upper()

        # Name side: Trial Summary Parameter Test Name
        tsparm_df = ct[
            ct["Codelist Name"] == "Trial Summary Parameter Test Name"
        ][["Code", "Submission Value"]].copy()

        tsparm_df = tsparm_df.rename(columns={
            "Submission Value": "TSPARM"
        })

        # 以 Code 配對
        tsparm_lookup_df = tsparmcd_df.merge(
            tsparm_df,
            on="Code",
            how="inner"
        ).drop_duplicates(subset=["TSPARMCD"])

        # 轉成 dict：TSPARMCD -> TSPARM
        tsparm_lookup = dict(
            zip(tsparm_lookup_df["TSPARMCD"], tsparm_lookup_df["TSPARM"])
        )

    # ----------------------------------------
    # TS（展開 + 自動填值）
    # ----------------------------------------
    ts_rows = []

    for i, tsparmcd in enumerate(tsparmcd_list, start=1):

        tsparm = tsparm_lookup.get(tsparmcd, "")
        tsval = tsval_map.get(tsparmcd, "")
        tsvcdref, tsvcdver = refver_map.get(tsparmcd, ("", ""))

        ts_rows.append({
            "STUDYID": protocol_no,
            "DOMAIN": "TS",
            "TSSEQ": "1",
            "TSPARMCD": tsparmcd,
            "TSPARM": tsparm,
            "TSVAL": tsval,
            "TSVALCD": "",
            "TSVCDREF": tsvcdref,
            "TSVCDVER": tsvcdver
        })

    ts_df = pd.DataFrame(ts_rows)

    # ----------------------------------------
    # TV（用 SoA visit）
    # ----------------------------------------
    tv_cols = [
        "STUDYID","DOMAIN","VISITNUM","VISIT",
        "VISITDY","ARMCD","ARM","TVSTRL","TVENRL"
    ]

    tv_rows = []

    if unique_visit_df is not None and not unique_visit_df.empty:

        df = unique_visit_df.copy()

        # 保底
        if "Visit" not in df.columns:
            df["Visit"] = ""

        if "Visit_order" not in df.columns:
            df["Visit_order"] = range(1, len(df) + 1)

        df["Visit"] = df["Visit"].astype(str).str.strip()
        df = df[df["Visit"] != ""]

        df = df.sort_values("Visit_order").reset_index(drop=True)

        for i, row in df.iterrows():
            tv_rows.append({
                "STUDYID": protocol_no,
                "DOMAIN": "TV",
                "VISITNUM": "",
                "VISIT": row["Visit"],
                "VISITDY": "",
                "ARMCD": "",
                "ARM": "",
                "TVSTRL": "",
                "TVENRL": ""
            })

    tv_df = pd.DataFrame(tv_rows, columns=tv_cols)


    return {
        "TA": ta_df,
        "TE": te_df,
        "TI": ti_df,
        "TS": ts_df,
        "TV": tv_df
    }
    # End=========================================================


# ----------------------------------------
# Helper小工具
# ----------------------------------------
def get_paired_variables(variable):
    """
    給一個 variable name，回傳應保留的 paired variables
    """
    variable = str(variable or "").strip().upper()
    paired = set()

    if variable.endswith("DTC"):
        paired.add(variable[:-3] + "DY")

    if variable.endswith("STDTC"):
        paired.add(variable[:-5] + "STDY")

    if variable.endswith("ENDTC"):
        paired.add(variable[:-5] + "ENDY")

    if variable.endswith("STRTPT"):
        paired.add(variable[:-6] + "STTPT")

    if variable.endswith("ENRTPT"):
        paired.add(variable[:-6] + "ENTPT")

    if variable == "VISITNUM":
        paired.update(["VISIT", "VISITDY"])

    if variable.endswith("TPT"):
        paired.add(variable + "NUM")

    if variable.endswith("ORRES"):
        base = variable[:-5]
        paired.update([
            base + "STRESC",
            base + "STRESN",
            base + "STAT"
        ])

    if variable.endswith("ORRESU"):
        base = variable[:-6]
        paired.add(base + "STRESU")

    return paired
    # End=========================================================



def expand_suppqual_variables(config_df, target_supp_datasets):
    """
    config 只有 SUPPQUAL 時，複製成 SUPPxx variables
    """
    if config_df is None or config_df.empty:
        return config_df.copy()

    cfg = config_df.copy()

    if "Dataset" not in cfg.columns:
        return cfg

    cfg["Dataset"] = cfg["Dataset"].astype(str).str.upper().str.strip()

    suppqual_rows = cfg[cfg["Dataset"] == "SUPPQUAL"].copy()

    if suppqual_rows.empty:
        return cfg

    expanded_parts = [cfg[cfg["Dataset"] != "SUPPQUAL"].copy()]

    for ds in sorted(set(target_supp_datasets)):
        ds = str(ds).strip().upper()
        if not ds.startswith("SUPP"):
            continue

        dup = suppqual_rows.copy()
        dup["Dataset"] = ds

        if "Dataset Label" in dup.columns:
            base_domain = ds[4:]   # SUPPAE -> AE
            dup["Dataset Label"] = f"Supplemental Qualifiers for {base_domain}"

        expanded_parts.append(dup)

    out = pd.concat(expanded_parts, ignore_index=True)
    out = out.drop_duplicates()

    return out.reset_index(drop=True)
    # End=========================================================


def join_unique(series):
    vals = [
        str(x).strip()
        for x in series.fillna("").astype(str).tolist()
        if str(x).strip() != ""
    ]
    vals = list(dict.fromkeys(vals))  # 保序去重
    return "; ".join(vals)
    # End=========================================================



def apply_codelist_rules(merged_df):
    """
    只在 Codelist 原本為空時，依規則補值
    """

    df = merged_df.copy()

    # 保底
    if "Dataset" not in df.columns:
        df["Dataset"] = ""
    if "Variable" not in df.columns:
        df["Variable"] = ""
    if "CT Code" in df.columns:
        df["Codelist"] = df["CT Code"]
    else:
        df["Codelist"] = ""

    df["Dataset"] = df["Dataset"].astype(str).str.upper().str.strip()
    df["Variable"] = df["Variable"].astype(str).str.upper().str.strip()
    df["Codelist"] = df["Codelist"].fillna("").astype(str).str.strip()

    # 只處理目前還沒有 Codelist 的列
    missing_mask = df["Codelist"] == ""

    # -------------------------------------------------
    # Rule 1: Codelist = Variable
    # - DOMAIN
    # - CO.RDOMAIN
    # - XXTEST
    # - XXTESTCD
    # -------------------------------------------------
    rule_var_as_codelist = (
        (df["Variable"] == "DOMAIN") |
        (df["Variable"] == "AEREL") |
        ((df["Dataset"] == "CO") & (df["Variable"] == "RDOMAIN")) |
        (df["Variable"].str.endswith("TEST")) |
        (df["Variable"].str.endswith("TESTCD"))
    )

    df.loc[missing_mask & rule_var_as_codelist, "Codelist"] = df.loc[
        missing_mask & rule_var_as_codelist, "Variable"
    ]

    # -------------------------------------------------
    # Rule 2: AE dictionary fields -> AEDICT_F
    # -------------------------------------------------
    ae_dict_vars = {
        "AELLT", "AELLTCD", "AEDECOD", "AEPTCD",
        "AEHLT", "AEHLTCD", "AEHLGT", "AEHLGTCD",
        "AEBODSYS", "AEBDSYCD", "AESOC", "AESOCCD"
    }

    mask_ae_dict = missing_mask & df["Variable"].isin(ae_dict_vars)
    df.loc[mask_ae_dict, "Codelist"] = "AEDICT_F"

    # -------------------------------------------------
    # Rule 3: ARMCD / ACTARMCD -> ARMCD
    # -------------------------------------------------
    mask_armcd = missing_mask & df["Variable"].isin(["ARMCD", "ACTARMCD"])
    df.loc[mask_armcd, "Codelist"] = "ARMCD"

    # -------------------------------------------------
    # Rule 4: ARM / ACTARM -> ARM
    # -------------------------------------------------
    mask_arm = missing_mask & df["Variable"].isin(["ARM", "ACTARM"])
    df.loc[mask_arm, "Codelist"] = "ARM"

    # -------------------------------------------------
    # Rule 5: COUNTRY -> ISO3166
    # -------------------------------------------------
    mask_country = missing_mask & (df["Variable"] == "COUNTRY")
    df.loc[mask_country, "Codelist"] = "ISO3166"

    # -------------------------------------------------
    # Rule 6: SUPP-- 的 RDOMAIN -> DOMAIN_{XX}
    # -------------------------------------------------
    mask_supp_rdomain = (
        missing_mask &
        df["Dataset"].str.startswith("SUPP") &
        (df["Variable"] == "RDOMAIN")
    )

    df.loc[mask_supp_rdomain, "Codelist"] = (
        "DOMAIN_" + df.loc[mask_supp_rdomain, "Dataset"].str.replace("SUPP", "", regex=False)
    )

    
    # -------------------------------------------------
    # Rule 7: 跨 domain Codelist 拆分
    # -------------------------------------------------

    cross_domain_targets = {"DOMAIN", "FREQ", "LOC", "METHOD", "NRIND", "UNIT", "ROUTE"}

    # 統一格式
    df["Dataset"] = df["Dataset"].astype(str).str.upper().str.strip()
    df["Codelist"] = df["Codelist"].astype(str).str.upper().str.strip()

    # 找每個 Codelist 對應到的 domain 數量
    codelist_domain_count = (
        df.groupby("Codelist")["Dataset"]
        .nunique()
    )

    # 找出跨 domain 的 codelist
    multi_domain_codelists = set(
        codelist_domain_count[codelist_domain_count > 1].index
    )

    # 只針對指定清單 + 跨domain的處理
    mask = df["Codelist"].isin(cross_domain_targets) & df["Codelist"].isin(multi_domain_codelists)

    df.loc[mask, "Codelist"] = (
        df.loc[mask, "Codelist"] + "_" + df.loc[mask, "Dataset"]
    )


    # -------------------------------------------------
    # Rule 8: STENRF 特殊規則（XXSTRTPT / XXENRTPT）
    # -------------------------------------------------

    df["Variable"] = df["Variable"].astype(str).str.upper().str.strip()
    df["Dataset"] = df["Dataset"].astype(str).str.upper().str.strip()

    mask_strtpt = df["Variable"].str.endswith("STRTPT")
    mask_enrtpt = df["Variable"].str.endswith("ENRTPT")

    # STRTPT → STENRF_{XX}_START
    df.loc[mask_strtpt, "Codelist"] = (
        "STENRF_" +
        df.loc[mask_strtpt, "Variable"].str.replace("STRTPT", "", regex=False) +
        "_START"
    )

    # ENRTPT → STENRF_{XX}_END
    df.loc[mask_enrtpt, "Codelist"] = (
        "STENRF_" +
        df.loc[mask_enrtpt, "Variable"].str.replace("ENRTPT", "", regex=False) +
        "_END"
    )

    return df
    # End=========================================================




def apply_origin_rules(df):
    """
    只在非 Collected 時調整 Origin / Source
    """

    df = df.copy()

    # 保底欄位
    for col in ["Dataset", "Variable", "Origin", "Source", "Pages", "Method", "Codelist"]:
        if col not in df.columns:
            df[col] = ""

    df["Dataset"] = df["Dataset"].astype(str).str.upper().str.strip()
    df["Variable"] = df["Variable"].astype(str).str.upper().str.strip()
    df["Origin"] = df["Origin"].astype(str).str.strip()
    df["Codelist"] = df["Codelist"].astype(str).str.upper().str.strip()

    # -------------------------------------------------
    # 0. 只處理非 Collected
    # -------------------------------------------------
    mask_non_collected = df["Origin"].str.upper() != "COLLECTED"

    # -------------------------------------------------
    # 1. Codelist = AEDICT_F -> Assigned + Vendor
    # -------------------------------------------------
    mask_aedict = mask_non_collected & (df["Codelist"] == "AEDICT_F")
    df.loc[mask_aedict, "Origin"] = "Assigned"
    df.loc[mask_aedict, "Source"] = "Vendor"

    mask_epoch = mask_non_collected & (df["Variable"] == "EPOCH") & (df["Dataset"] == "TA")
    df.loc[mask_epoch, "Origin"] = "Assigned"

    mask_visitdy = mask_non_collected & (df["Variable"] == "VISITDY")
    df.loc[mask_visitdy, "Origin"] = "Protocol"
    df.loc[mask_visitdy, "Source"] = "Sponsor"

    # -------------------------------------------------
    # 2. 強制覆寫 Origin（僅限非 Collected 且非 AEDICT_F）
    # -------------------------------------------------
    mask_target = mask_non_collected & ~mask_aedict & ~mask_epoch & ~mask_visitdy
    
    mask_protocol_vars = mask_target & (
        df["Variable"].isin(["STUDYID", "ECTRT", "ECDOSE", "ECDOSU", "ECDOSFRM", "EXTRT", "EXDOSE", "EXDOSU", "EXDOSFRM"])
    )
    df.loc[mask_protocol_vars, "Origin"] = "Protocol"

    
    assigned_patterns = ["TPTNUM", "CAT", "TEST", "TESTCD"]
    mask_assigned_vars = (
        mask_target &
        ( df["Variable"].str.endswith(tuple(assigned_patterns)) |
         df["Variable"].isin(["DOMAIN", "RDOMAIN", "VISITNUM", "VISIT", "IDVAR", "IDVARVAL", "QNAM", "QLABEL", "QORIG", "QEVAL", "COREF", "COEVAL", 
                              "AGEU", "ARMCD", "ARM", "ACTARMCD", "ACTARM", "ARMNRS", "ACTARMUD", "ETCD", "SVPRESP", "TAETORD", "TSPARMCD", "TSPARM", "TSVALCD", "TSVCDREF", "TSVCDVER"])
        )
    )
    df.loc[mask_assigned_vars, "Origin"] = "Assigned"


    derived_patterns = ["SEQ", "DY", "STDY", "ENDY", "ENTPT", "STTPT", "STRESC", "STRESN", "STRESU", "STAT", "LOBXFL"]
    mask_derived_vars = (
        mask_target &
        ( df["Variable"].str.endswith(tuple(derived_patterns)) |
         df["Variable"].isin(["USUBJID", "EPOCH", "RFSTDTC", "RFENDTC", "RFXSTDTC", "RFXENDTC", "RFPENDTC", "DTHFL", "DSDTC", "SESTDTC", "SEENDTC"])
        )
    )
    df.loc[mask_derived_vars, "Origin"] = "Derived"


    # -------------------------------------------------
    # 3. Source 規則
    #    - AEDICT_F 已經先設成 Vendor
    #    - 其他 Protocol / Derived / Assigned -> Sponsor
    # -------------------------------------------------
    mask_protocol = mask_target & (df["Origin"].str.upper() == "PROTOCOL")
    mask_derived = mask_target & (df["Origin"].str.upper() == "DERIVED")
    mask_assigned = mask_target & (df["Origin"].str.upper() == "ASSIGNED")

    df.loc[mask_protocol, "Source"] = "Sponsor"
    df.loc[mask_derived, "Source"] = "Sponsor"
    df.loc[mask_assigned, "Source"] = "Sponsor"

    
   # -------------------------------------------------
    # Rule X: LB fallback（Origin 還是空 → Collected + Vendor）
    # -------------------------------------------------

    mask_lb_fallback = (
        df["Dataset"].str.startswith("LB") &
        (df["Origin"].astype(str).str.strip() == "")
    )

    df.loc[mask_lb_fallback, "Origin"] = "Collected"
    df.loc[mask_lb_fallback, "Source"] = "Vendor"

    return df
    # End=========================================================




def apply_method_rules(df):

    df = df.copy()

    # 保底欄位
    for col in ["Variable", "Dataset", "Method", "Comment"]:
        if col not in df.columns:
            df[col] = ""

    df["Variable"] = df["Variable"].astype(str).str.upper().str.strip()
    df["Dataset"] = df["Dataset"].astype(str).str.upper().str.strip()
    df["Method"] = df["Method"].fillna("").astype(str)
    df["Comment"] = df["Comment"].fillna("").astype(str)

    # -------------------------------------------------
    # Method rules
    # -------------------------------------------------

    # 1. USUBJID
    df.loc[df["Variable"] == "USUBJID", "Method"] = \
        "Concatenation of STUDYID-SITEID-SUBJID"

    # 2. TSSEQ
    df.loc[df["Variable"] == "TSSEQ", "Method"] = \
        "Equal to sequential number identifying records within each TSPARMCD in the domain"

    # 3. {XX}SEQ（除了 TS）
    mask_seq = df["Variable"].str.endswith("SEQ") & (df["Dataset"] != "TS")
    df.loc[mask_seq, "Method"] = \
        "Equal to sequential number identifying records within each USUBJID sorted by key variables in the domain"

    # 4. {XX}{ST/EN}DY
    mask_stdy = df["Variable"].str.endswith("STDY")
    mask_endy = df["Variable"].str.endswith("ENDY")
    mask_dy = (
        ( df["Variable"].str.endswith("DY") & (df["Variable"] != "VISITDY") ) &
        ~mask_stdy &
        ~mask_endy
    )

    # STDY → STDTC
    df.loc[mask_stdy, "Method"] = df.loc[mask_stdy, "Variable"].apply(
        lambda x: (
            f"Equal to {x.replace('STDY','STDTC')} - DM.RFSTDTC + 1 if {x.replace('STDY','STDTC')} is on or after DM.RFSTDTC; "
            f"equal to {x.replace('STDY','STDTC')} - DM.RFSTDTC if {x.replace('STDY','STDTC')} precedes DM.RFSTDTC"
        )
    )

    # ENDY → ENDTC
    df.loc[mask_endy, "Method"] = df.loc[mask_endy, "Variable"].apply(
        lambda x: (
            f"Equal to {x.replace('ENDY','ENDTC')} - DM.RFSTDTC + 1 if {x.replace('ENDY','ENDTC')} is on or after DM.RFSTDTC; "
            f"equal to {x.replace('ENDY','ENDTC')} - DM.RFSTDTC if {x.replace('ENDY','ENDTC')} precedes DM.RFSTDTC"
        )
    )

    # DY → DTC（排除 STDY / ENDY）
    df.loc[mask_dy, "Method"] = df.loc[mask_dy, "Variable"].apply(
        lambda x: (
            f"Equal to {x.replace('DY','DTC')} - DM.RFSTDTC + 1 if {x.replace('DY','DTC')} is on or after DM.RFSTDTC; "
            f"equal to {x.replace('DY','DTC')} - DM.RFSTDTC if {x.replace('DY','DTC')} precedes DM.RFSTDTC"
        )
    )


    # 5. RFSTDTC
    df.loc[df["Variable"] == "RFSTDTC", "Method"] = \
        "Equal to date/time of first exposure to study treatment (earliest EXSTDTC)"

    # 6. RFENDTC
    df.loc[df["Variable"] == "RFENDTC", "Method"] = \
        "Equal to date/time of last exposure to study treatment (latest EXENDTC)"

    # 7. RFXSTDTC / RFXENDTC
    df.loc[df["Variable"] == "RFXSTDTC", "Method"] = "Equal to RFSTDTC"
    df.loc[df["Variable"] == "RFXENDTC", "Method"] = "Equal to RFENDTC"

    # 8. RFPENDTC
    df.loc[df["Variable"] == "RFPENDTC", "Method"] = \
        "Equal to the last known date during the study"

    # 9. DTHFL
    df.loc[df["Variable"] == "DTHFL", "Method"] = \
        'Set to "Y" if DTHDTC is populated'

    # 10. DSDTC
    df.loc[df["Variable"] == "DSDTC", "Method"] = "Equal to DSSTDTC"


    # -------------------------------------------------
    # STRES 系列（動態連動）
    # -------------------------------------------------
    # STRESC → ORRES
    mask_stresc = df["Variable"].str.endswith("STRESC")
    df.loc[mask_stresc, "Method"] = df.loc[mask_stresc, "Variable"].apply(
        lambda x: f"Equal to {x.replace('STRESC','ORRES')}"
    )

    # STRESN → STRESC
    mask_stresn = df["Variable"].str.endswith("STRESN")
    df.loc[mask_stresn, "Method"] = df.loc[mask_stresn, "Variable"].apply(
        lambda x: (
            f"Equal to numeric value of {x.replace('STRESN','STRESC')} "
            f"if {x.replace('STRESN','STRESC')} contains numeric data"
        )
    )

    # STRESU → ORRESU
    mask_stresu = df["Variable"].str.endswith("STRESU")
    df.loc[mask_stresu, "Method"] = df.loc[mask_stresu, "Variable"].apply(
        lambda x: f"Equal to {x.replace('STRESU','ORRESU')}"
    )

    # STAT → ORRES null
    mask_stat = df["Variable"].str.endswith("STAT")
    df.loc[mask_stat, "Method"] = df.loc[mask_stat, "Variable"].apply(
        lambda x: f'Equal to "NOT DONE" if {x.replace("STAT","ORRES")} is null'
    )

    # LOBXFL
    mask_lobxfl = df["Variable"].str.endswith("LOBXFL")
    df.loc[mask_lobxfl, "Method"] = \
        'Equal to "Y" for last record with non-missing value on or before DM.RFSTDTC; null otherwise'

    # -------------------------------------------------
    # Comment rules
    # -------------------------------------------------

    # IDVAR
    df.loc[df["Variable"] == "IDVAR", "Comment"] = \
        "Name of the variables for related records, such as --SEQ, VISIT or --DTC"

    # IDVARVAL
    df.loc[df["Variable"] == "IDVARVAL", "Comment"] = \
        "Value of identifying variable described in IDVAR"

    # VISITNUM（除了 TV）
    df.loc[
        (df["Variable"] == "VISITNUM") & (df["Dataset"] != "TV"),
        "Comment"
    ] = "Assigned from the TV domain based on VISIT"

    # TA.EPOCH
    df.loc[
        (df["Dataset"] == "TA") & (df["Variable"] == "EPOCH"),
        "Comment"
    ] = "Assigned based on protocol design"

    return df
    # End=========================================================



def build_define_sheet(protocol_no, protocol_title, sdtm_version):
    std_ver = sdtm_version.replace("Version", "").strip()
    define_records = [
        ["StudyName", protocol_no],
        ["StudyDescription", protocol_title],
        ["ProtocolName", protocol_no],
        ["StandardName", "SDTM-IG"],
        ["StandardVersion", std_ver],
        ["Language", "en"]
    ]

    define_df = pd.DataFrame(define_records, columns=["Attribute", "Value"])

    return define_df
    # End=========================================================



def build_datasets_from_variables(variables_df, config_df, sdtm_version):

    # 1. Dataset list
    dataset_df = (
        variables_df[["Dataset"]]
        .dropna()
        .drop_duplicates()
        .reset_index(drop=True)
    )

    dataset_df["Dataset"] = dataset_df["Dataset"].astype(str).str.upper().str.strip()


    # 2. Config metadata
    cfg = config_df.copy()
    
    cfg = cfg.rename(columns={
        "Dataset Label": "Label"
    })

    cfg["Dataset"] = cfg["Dataset"].astype(str).str.upper().str.strip()

    config_cols = ["Dataset", "Label", "Class", "Structure", "Key Variables"]
    cfg_meta = cfg[[c for c in config_cols if c in cfg.columns]]

    cfg_meta = (
        cfg_meta
        .groupby("Dataset", as_index=False)
        .agg({
            "Label": "first",
            "Class": "first",
            "Structure": "first",
            "Key Variables": "first"
        })
    )   

    # 拿 SUPPQUAL template
    suppqual_meta = cfg_meta[cfg_meta["Dataset"] == "SUPPQUAL"]


    # 3. Merge
    dataset_df = dataset_df.merge(
        cfg_meta,
        on="Dataset",
        how="left"
    )

    # 4. SUPPxx 補值（從 SUPPQUAL 套）
    if not suppqual_meta.empty:

        supp_mask = dataset_df["Dataset"].str.startswith("SUPP")

        for col in ["Label", "Class", "Structure", "Key Variables"]:
            if col in suppqual_meta.columns:
                dataset_df.loc[supp_mask, col] = dataset_df.loc[supp_mask, col].fillna(
                    suppqual_meta.iloc[0][col]
                )

        # Label 特別處理（較好）
        dataset_df.loc[supp_mask, "Label"] = dataset_df.loc[supp_mask, "Dataset"].apply(
            lambda x: f"Supplemental Qualifiers for {x.replace('SUPP','')}"
        )


    # 5. Standard 欄位
    std_ver = str(sdtm_version).upper().replace("VERSION", "").strip()
    dataset_df["Standard"] = f"SDTMIG {std_ver}"


    # 6. 排序
    final_cols = ["Dataset", "Label", "Class", "Structure", "Key Variables", "Standard"]
    dataset_df = dataset_df[[c for c in final_cols if c in dataset_df.columns]]

    dataset_df = dataset_df.sort_values("Dataset").reset_index(drop=True)

    return dataset_df
    # End=========================================================



def build_variables_sheet(detail_df, config_df, td_dict=None):
    """
    Variables Sheet
    來源：
      1. Step 1 detail_df（CRF -> SDTM Variable Mapping）
      2. 5T variable structure
      3. config_df 補 metadata
    """
    
    # 輸出欄位：
    final_cols = [
        "Order", "Dataset", "Variable", "Label", "Data Type",
        "CT Code", "Codelist", "Origin", "Source", "Pages", "Method", "Comment",
        "CRF Dataset", "CRF Variable",
    ]


    # -------------------------------------------------
    # 0. 保底
    # -------------------------------------------------
    if config_df is None or config_df.empty:
        return pd.DataFrame(columns=final_cols)

    cfg = config_df.copy()

    # config 欄位保底
    for col in ["Dataset", "Variable"]:
        if col not in cfg.columns:
            raise ValueError(f"config_df 缺少必要欄位: {col}")

    cfg["Dataset"] = cfg["Dataset"].astype(str).str.upper().str.strip()
    cfg["Variable"] = cfg["Variable"].astype(str).str.upper().str.strip()

    if "Core" in cfg.columns:
        cfg["Core"] = cfg["Core"].astype(str).str.upper().str.strip()


    # -------------------------------------------------
    # 1. 從 SDTM Variable Mapping - Detail
    # -------------------------------------------------
    detail_rows = []

    if detail_df is not None and not detail_df.empty:

        work = detail_df.copy()

        # 欄位保底
        for col in ["SDTM Domain", "SDTM Variable"]:
            if col not in work.columns:
                raise ValueError(f"detail_df 缺少必要欄位: {col}")

        for _, row in work.iterrows():

            dataset = str(row.get("SDTM Domain", "")).strip().upper()
            variable = str(row.get("SDTM Variable", "")).strip().upper()
            crf_dataset = str(row.get("CRF Dataset", "")).strip()
            crf_variable = str(row.get("CRF Variable", "")).strip()
            assign_value = str(row.get("Assign Value", "")).strip()

            if not dataset or not variable:
                continue

            # Origin / Method / Source 先用穩定邏輯
            if assign_value:
                origin = "Assigned"
                source = "Sponsor"
            else:
                origin = "Collected"
                source = "Investigator"

            detail_rows.append({
                "Dataset": dataset,
                "Variable": variable,
                "Origin": origin,
                "Source": source,
                "Pages": "",
                "Method": "",
                "Comment": "",
                "CRF Dataset": crf_dataset,
                "CRF Variable": crf_variable
            })

    detail_variables_df = pd.DataFrame(detail_rows)

    if not detail_variables_df.empty:
        detail_variables_df = (
            detail_variables_df
            .groupby(["Dataset", "Variable"], as_index=False)
            .agg({
                "Origin": "first",
                "Source": "first",
                "Pages": "first",
                "Method": "first",
                "Comment": "first",
                "CRF Dataset": join_unique,
                "CRF Variable": join_unique
            })
            .reset_index(drop=True)
        )

    # -------------------------------------------------
    # 2. 加入 5T variables
    # -------------------------------------------------
    td_rows = []
    
    if td_dict is not None:
        for dataset, df in td_dict.items():
            if df is None or df.empty:
                continue

            for col in df.columns:
                td_rows.append({
                    "Dataset": dataset,
                    "Variable": col,
                    "Origin": "Protocol",
                    "Source": "Sponsor",
                    "Pages": "",
                    "Method": "",
                    "Comment": "",
                    "CRF Dataset": "",
                    "CRF Variable": ""
                })

    td_variables_df = pd.DataFrame(td_rows)

    # -------------------------------------------------
    # 3. 合併來源 variables（Step1 + 5T）
    # -------------------------------------------------
    source_variables_df = pd.concat(
        [detail_variables_df, td_variables_df],
        ignore_index=True
    )
    if source_variables_df.empty:
        source_variables_df = pd.DataFrame(columns=[
            "Dataset", "Variable", "Origin", "Source", "Pages", "Method", "Comment", "CRF Dataset", "CRF Variable"
        ])
    else:
        source_variables_df = source_variables_df.drop_duplicates(
            subset=["Dataset", "Variable"],
            keep="first"
        ).reset_index(drop=True)

    # -------------------------------------------------
    # 4. 先定義 target datasets
    # -------------------------------------------------
    target_datasets = set(source_variables_df["Dataset"].dropna().astype(str).str.upper().tolist())
    target_datasets.update(["SV", "SE"])  # 強制留SV/SE

    # 先把 config 限縮到 target datasets + SUPPQUAL
    cfg_target = cfg[
        cfg["Dataset"].isin(target_datasets) | (cfg["Dataset"] == "SUPPQUAL")
    ].copy()


   # -------------------------------------------------
    # 5. 先展開 SUPPQUAL -> SUPPxx
    # -------------------------------------------------
    target_supp_datasets = sorted(
        source_variables_df["Dataset"][
            source_variables_df["Dataset"].astype(str).str.upper().str.startswith("SUPP")
        ]
        .dropna()
        .astype(str)
        .str.upper()
        .unique()
        .tolist()
    )

    expanded_cfg = expand_suppqual_variables(cfg_target, target_supp_datasets)
    # SUPPQUAL 本身不呈現
    expanded_cfg = expanded_cfg[expanded_cfg["Dataset"] != "SUPPQUAL"].copy()

    
    # -------------------------------------------------
    # 6. Config 額外保留規則
    # -------------------------------------------------
    cfg_keep_mask = pd.Series(False, index=expanded_cfg.index)

    # SV / SE 保留 -> 但仍只看 target datasets 內的 expanded_cfg
    #cfg_keep_mask = cfg_keep_mask | expanded_cfg["Dataset"].isin(["SV", "SE"])

    # Core = REQUIRED / EXPECTED
    if "Core" in expanded_cfg.columns:
        cfg_keep_mask = cfg_keep_mask | expanded_cfg["Core"].astype(str).str.upper().isin(["REQUIRED", "EXPECTED"])

    # Variable = EPOCH
    cfg_keep_mask = cfg_keep_mask | (expanded_cfg["Variable"] == "EPOCH")

    # 強制留下的 variables
    existing_domains = set(
        source_variables_df["Dataset"]
        .dropna()
        .astype(str)
        .str.upper()
        .tolist()
    )

    force_keep_map = {}
    
    if "CO" in existing_domains:
        force_keep_map["CO"] = {"RDOMAIN", "IDVAR", "IDVARVAL", "COREF", "COEVAL"}

    if "DS" in existing_domains:
        force_keep_map["DS"] = {"DSDTC"}


    force_keep_mask = pd.Series(False, index=expanded_cfg.index)
    for ds, vars_set in force_keep_map.items():
        force_keep_mask = force_keep_mask | (
            (expanded_cfg["Dataset"] == ds) &
            (expanded_cfg["Variable"].isin(vars_set))
        )
        
    cfg_keep_mask = cfg_keep_mask | force_keep_mask


    cfg_keep_df = expanded_cfg.loc[cfg_keep_mask, ["Dataset", "Variable"]].drop_duplicates().copy()
    if not cfg_keep_df.empty:
        cfg_keep_df["Origin"] = ""
        cfg_keep_df["Source"] = ""
        cfg_keep_df["Pages"] = ""
        cfg_keep_df["Method"] = ""
        cfg_keep_df["Comment"] = ""
        cfg_keep_df["CRF Dataset"] = ""
        cfg_keep_df["CRF Variable"] = ""

    # Paired variables
    existing_pairs_source = pd.concat(
        [source_variables_df[["Dataset", "Variable"]], cfg_keep_df[["Dataset", "Variable"]]],
        ignore_index=True
    ).drop_duplicates()

    pair_rows = []
    existing_pairs_set = set(
        zip(existing_pairs_source["Dataset"], existing_pairs_source["Variable"])
    )

    for _, row in existing_pairs_source.iterrows():
        dataset = row["Dataset"]
        variable = row["Variable"]

        for paired_var in get_paired_variables(variable):
            pair_rows.append({
                "Dataset": dataset,
                "Variable": paired_var,
                "Origin": "",
                "Source": "",
                "Pages": "",
                "Method": "",
                "Comment": "",
                "CRF Dataset": "",
                "CRF Variable": ""
            })

    pair_df = pd.DataFrame(pair_rows)
    if not pair_df.empty:
        pair_df = pair_df.drop_duplicates(subset=["Dataset", "Variable"], keep="first")

        # 只保留 expanded_cfg 裡真的存在的 paired vars
        cfg_pair_key = set(zip(expanded_cfg["Dataset"], expanded_cfg["Variable"]))
        pair_df = pair_df[
            pair_df.apply(lambda r: (r["Dataset"], r["Variable"]) in cfg_pair_key, axis=1)
        ].reset_index(drop=True)

    # -------------------------------------------------
    # 7. 合併全部 variables universe
    # -------------------------------------------------
    variables_universe = pd.concat(
        [source_variables_df, cfg_keep_df, pair_df],
        ignore_index=True
    )

    if variables_universe.empty:
        return pd.DataFrame(columns=final_cols)

    variables_universe = variables_universe.drop_duplicates(
        subset=["Dataset", "Variable"],
        keep="first"
    ).reset_index(drop=True)



    # -------------------------------------------------
    # 8. merge config metadata
    # -------------------------------------------------
    cfg_cols = [c for c in [
        "Dataset", "Variable", "VarNum", "Variable Label", "Data Type", "CT Code"
    ] if c in expanded_cfg.columns]

    cfg_meta = expanded_cfg[cfg_cols].drop_duplicates(subset=["Dataset", "Variable"], keep="first")

    merged = variables_universe.merge(
        cfg_meta,
        how="left",
        on=["Dataset", "Variable"]
    )

    merged = merged.rename(columns={
        "Variable Label": "Label",
        "CTcode": "CT Code"
    })

    merged = apply_codelist_rules(merged)
    merged = apply_origin_rules(merged)
    merged = apply_method_rules(merged)
    
    # Data Type 轉換
    if "Data Type" in merged.columns:
        merged["Data Type"] = merged["Data Type"].apply(
            lambda x: "integer" if pd.notna(x) and int(float(x)) == 1
            else ("text" if pd.notna(x) and int(float(x)) == 2
                  else str(x).strip())

        )


    # -------------------------------------------------
    # 9. 保底欄位
    # -------------------------------------------------
    for col in final_cols:
        if col not in merged.columns:
            merged[col] = ""

    # -------------------------------------------------
    # 10. Order：各 dataset 內按 Variable 排序後重編
    # -------------------------------------------------
    out_parts = []

    for dataset, grp in merged.groupby("Dataset", dropna=False):
        grp = grp.copy()

        grp["VarNum"] = pd.to_numeric(grp["VarNum"], errors="coerce")
        grp = grp.sort_values(by=["VarNum"]).reset_index(drop=True)
        grp["Order"] = range(1, len(grp) + 1)

        out_parts.append(grp)

    final_df = pd.concat(out_parts, ignore_index=True)

    # -------------------------------------------------
    # 11. 最終輸出
    # -------------------------------------------------
    final_df = final_df[final_cols].copy()
    final_df = final_df.sort_values(by=["Dataset", "Order", "Variable"]).reset_index(drop=True)

    return final_df
    # End=========================================================




def build_codelist_sheet(variables_spec_df, ct_master_df=None, matched_ct_df=None, ct_mapping_df=None, ts_df=None, sdtm_ct=None):

    df = variables_spec_df.copy()

    # =================================================
    # 0. 保底欄位
    # =================================================
    for col in ["Dataset", "Variable", "Label", "CT Code", "Codelist"]:
        if col not in df.columns:
            df[col] = ""

    df["Dataset"] = df["Dataset"].fillna("").astype(str).str.strip().str.upper()
    df["Variable"] = df["Variable"].fillna("").astype(str).str.strip().str.upper()
    df["Label"] = df["Label"].fillna("").astype(str).str.strip()
    df["CT Code"] = df["CT Code"].fillna("").astype(str).str.strip().str.upper()
    df["Codelist"] = df["Codelist"].fillna("").astype(str).str.strip().str.upper()

    # 只保留有 Codelist 的，移除不要進 2.4 的
    df = df[
        (df["Codelist"] != "") &
        (~df["Codelist"].isin(["AEDICT_F", "ISO3166"]))
    ].copy()

    # =================================================
    # 針對 DOMAIN codelist，先排除 SUPPxx
    # 避免 DOMAIN_VS 被 SUPPVS 蓋掉
    # =================================================
    is_domain_codelist = df["Codelist"].fillna("").astype(str).str.upper().str.startswith("DOMAIN")
    is_supp_dataset = df["Dataset"].fillna("").astype(str).str.upper().str.startswith("SUPP")

    df = df[~(is_domain_codelist & is_supp_dataset)].copy()


    # 排序後挑 representative row
    df = df.sort_values(by=["Codelist", "Dataset", "Variable"]).reset_index(drop=True)

    codelist_df = (
        df[["Dataset", "Variable", "Label", "CT Code", "Codelist"]]
        .drop_duplicates(subset=["Codelist"], keep="first")
        .reset_index(drop=True)
    )

    # =================================================
    # 1. ID / ID_Temp
    # =================================================
    codelist_df["ID"] = codelist_df["Codelist"]

    codelist_df["ID_Temp"] = (
        codelist_df["Codelist"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
        .apply(lambda x: x.split("_")[0] if x else "")
    )

    # Y subset 要對到 NY codelist
    codelist_df.loc[
        codelist_df["ID"] == "Y",
        "ID_Temp"
    ] = "NY"

    codelist_df = codelist_df.sort_values("Codelist").reset_index(drop=True)

    # =================================================
    # 2. 準備 matched_df（情況1：CT Code -> CT Term）
    # =================================================
    if matched_ct_df is not None and not matched_ct_df.empty:
        matched_df = matched_ct_df.copy()
        matched_df.columns = [str(c).strip() for c in matched_df.columns]

        rename_map = {}
        for c in matched_df.columns:
            cu = c.upper().strip()
            if cu == "SDTM DOMAIN":
                rename_map[c] = "Dataset"
            elif cu == "SDTM VARIABLE":
                rename_map[c] = "Variable"

        matched_df = matched_df.rename(columns=rename_map)

        for col in ["Dataset", "Variable", "CT Code", "CT Term"]:
            if col not in matched_df.columns:
                matched_df[col] = ""

        matched_df["Dataset"] = matched_df["Dataset"].fillna("").astype(str).str.strip().str.upper()
        matched_df["Variable"] = matched_df["Variable"].fillna("").astype(str).str.strip().str.upper()
        matched_df["CT Code"] = matched_df["CT Code"].fillna("").astype(str).str.strip().str.upper()
        matched_df["CT Term"] = matched_df["CT Term"].fillna("").astype(str).str.strip()

    else:
        matched_df = pd.DataFrame(columns=["Dataset", "Variable", "CT Code", "CT Term"])

    # =================================================
    # 3. 準備 map_df（情況2：Assign / CRF Option / Original）
    # =================================================
    if ct_mapping_df is not None and not ct_mapping_df.empty:
        map_df = ct_mapping_df.copy()
        map_df.columns = [str(c).strip() for c in map_df.columns]

        rename_map = {}
        for c in map_df.columns:
            cu = c.upper().strip()
            if cu == "SDTM DOMAIN":
                rename_map[c] = "Dataset"
            elif cu == "SDTM VARIABLE":
                rename_map[c] = "Variable"

        map_df = map_df.rename(columns=rename_map)

        for col in ["Dataset", "Variable", "CT Code", "Assign Value", "CRF Option Value", "Original Value"]:
            if col not in map_df.columns:
                map_df[col] = ""

        map_df["Dataset"] = map_df["Dataset"].fillna("").astype(str).str.strip().str.upper()
        map_df["Variable"] = map_df["Variable"].fillna("").astype(str).str.strip().str.upper()
        map_df["CT Code"] = map_df["CT Code"].fillna("").astype(str).str.strip().str.upper()
        map_df["Assign Value"] = map_df["Assign Value"].fillna("").astype(str).str.strip()
        map_df["CRF Option Value"] = map_df["CRF Option Value"].fillna("").astype(str).str.strip()
        map_df["Original Value"] = map_df["Original Value"].fillna("").astype(str).str.strip()

    else:
        map_df = pd.DataFrame(
            columns=["Dataset", "Variable", "CT Code", "Assign Value", "CRF Option Value", "Original Value"]
        )

    # =================================================
    # 4. 準備 TS（特殊處理）
    # =================================================
    if ts_df is not None and not ts_df.empty:
        ts_work = ts_df.copy()

        for col in ["TSPARM", "TSPARMCD"]:
            if col not in ts_work.columns:
                ts_work[col] = ""

        ts_work["TSPARM"] = ts_work["TSPARM"].fillna("").astype(str).str.strip()
        ts_work["TSPARMCD"] = ts_work["TSPARMCD"].fillna("").astype(str).str.strip()

    else:
        ts_work = pd.DataFrame(columns=["TSPARM", "TSPARMCD"])

    # =================================================
    # 5. 第一層 merge（Main Codelist Code）
    # =================================================
    if ct_master_df is not None and not ct_master_df.empty:

        right = ct_master_df.copy()

        for col in ["Submission Value", "Codelist Code", "Code", "Codelist Name"]:
            if col not in right.columns:
                right[col] = ""

        right["Submission Value"] = (
            right["Submission Value"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )

        right["Codelist Code"] = (
            right["Codelist Code"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        right["Codelist Name"] = (
            right["Codelist Name"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        # Main code only
        right_main = right[right["Codelist Code"] == ""].copy()

        codelist_df = codelist_df.merge(
            right_main[
                ["Submission Value", "Code", "Codelist Name"]
            ].drop_duplicates(),
            left_on="ID_Temp",
            right_on="Submission Value",
            how="left"
        )

        codelist_df = codelist_df.rename(columns={
            "Code": "NCI Codelist Code",
            "Codelist Name": "Name"
        })
        
        codelist_df["Terminology"] = codelist_df["NCI Codelist Code"].apply(
            lambda x: f"SDTM {sdtm_ct}"
            if pd.notna(x) and str(x).strip() not in ["", "None", "nan"]
            else ""
        )

        codelist_df = codelist_df.drop(columns=["Submission Value"], errors="ignore")

    # =================================================
    # 6. Name 調整（不依賴第一層一定成功）
    # =================================================
    if "Name" not in codelist_df.columns:
        codelist_df["Name"] = ""

    codelist_df["Name"] = codelist_df["Name"].fillna("").astype(str).str.strip()

    # 沒有 merge 到 → 用 Label
    codelist_df.loc[
        codelist_df["Name"] == "",
        "Name"
    ] = codelist_df["Label"]

    # 特殊 override
    special_map = {
        "ARM": "Description of Arm",
        "ARMCD": "Arm Code",
        "Y": "No Yes Response - Y subset"
    }

    codelist_df["Name"] = codelist_df.apply(
        lambda row: special_map.get(row["ID"], row["Name"]),
        axis=1
    )

    # 跨 domain 的 codelist 顯示加上 domain
    mask = codelist_df["Codelist"].str.contains("_", na=False)
    codelist_df.loc[mask, "Name"] = (
        codelist_df.loc[mask, "Name"]
        + " (" + codelist_df.loc[mask, "Dataset"] + ")"
    )

    # =================================================
    # 7. 第二層：展開 Term
    # =================================================
    expanded_rows = []

    for _, row in codelist_df.iterrows():

        dataset = row["Dataset"]
        variable = row["Variable"]
        label = row["Label"]
        ct_code = str(row["CT Code"]).strip().upper()
        codelist = row["Codelist"]
        id_ = row["ID"]
        id_temp = row["ID_Temp"]
        name = row["Name"]
        nci_codelist_code = row.get("NCI Codelist Code", "")
        terminology = row.get("Terminology", "")
        
        terms = []

        # ---------------------------------------------
        # 情況 1：有 CT Code
        # 先抓 matched CT term
        # 再保留 unmatched original value
        # ---------------------------------------------
        if ct_code != "":
            subset = map_df[
                (map_df["Dataset"] == dataset) &
                (map_df["Variable"] == variable)
            ].copy()

            matched_subset = matched_df.loc[
                (matched_df["CT Code"] == ct_code) &
                (matched_df["Dataset"] == dataset) &
                (matched_df["Variable"] == variable)
            ].copy()

            matched_terms = (
                matched_subset["CT Term"]
                .dropna()
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .dropna()
                .drop_duplicates()
                .tolist()
            )

            original_terms = (
                subset["Original Value"]
                .dropna()
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .dropna()
                .drop_duplicates()
                .tolist()
            )

            if "Original Value Normalized" in matched_subset.columns:
                matched_original_norm = set(
                matched_subset["Original Value Normalized"]
                .dropna()
                .astype(str)
                .str.strip()
                .str.upper()
                .tolist()
            )
            else:
                matched_original_norm = set(
                    matched_subset["Original Value"]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .apply(normalize_text)
                    .tolist()
                )

            if id_ != "NY":
                unmatched_original_terms = [
                    t for t in original_terms
                    if normalize_text(t) not in matched_original_norm
                ]
            else:
                unmatched_original_terms = []
            
            # 保留 matched + unmatched
            terms = matched_terms + unmatched_original_terms

        # ---------------------------------------------
        # 情況 2：沒 CT Code → 優先 Assign Value，再 CRF Option，再 Original
        # ---------------------------------------------
        elif ct_code == "":
            subset = map_df[
                (map_df["Dataset"] == dataset) &
                (map_df["Variable"] == variable)
            ].copy()

            assign_terms = (
                subset["Assign Value"]
                .dropna()
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .dropna()
                .drop_duplicates()
                .tolist()
            )

            option_terms = (
                subset["CRF Option Value"]
                .dropna()
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .dropna()
                .drop_duplicates()
                .tolist()
            )

            if assign_terms:
                terms = assign_terms
            elif option_terms:
                terms = option_terms
            else:
                fallback_terms = (
                    subset["Original Value"]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .replace("", pd.NA)
                    .dropna()
                    .drop_duplicates()
                    .tolist()
                )

                terms = fallback_terms

        # ---------------------------------------------
        # 情況 3：特殊處理（最後 fallback）
        # ---------------------------------------------
        if not terms:

            if id_temp == "DOMAIN":
                terms = [dataset]

            elif id_temp == "TSPARM":
                terms = (
                    ts_work["TSPARM"]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .replace("", pd.NA)
                    .dropna()
                    .drop_duplicates()
                    .tolist()
                )

            elif id_temp == "TSPARMCD":
                terms = (
                    ts_work["TSPARMCD"]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .replace("", pd.NA)
                    .dropna()
                    .drop_duplicates()
                    .tolist()
                )

            elif id_temp == "ND":
                terms = ["NOT DONE"]

            elif id_ == "NY":
                terms = ["N", "Y"]

            elif id_ == "Y":
                terms = ["Y"]

        # 如果還是沒有，就至少保留一列空值
        if not terms:
            terms = [""]

        for term in terms:
            expanded_rows.append({
                "Dataset": dataset,
                "Variable": variable,
                "Label": label,
                "CT Code": ct_code,
                "Codelist": codelist,
                "ID": id_,
                "ID_Temp": id_temp,
                "Name": name,
                "NCI Codelist Code": nci_codelist_code,
                "Terminology": terminology,
                "Term": term
            })

    codelist_df = pd.DataFrame(expanded_rows)

    # =================================================
    # 8. Term → NCI Term Code
    # =================================================
    if ct_master_df is not None and not ct_master_df.empty:

        right = ct_master_df.copy()

        for col in ["Submission Value", "Code", "Codelist Code"]:
            if col not in right.columns:
                right[col] = ""

        right["Submission Value"] = (
            right["Submission Value"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )

        right["Codelist Code"] = (
            right["Codelist Code"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        codelist_df["Term_norm"] = (
            codelist_df["Term"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )

        right_term = right[right["Codelist Code"] != ""].copy()

        codelist_df = codelist_df.merge(
            right_term[
                ["Submission Value", "Code", "Codelist Code"]
            ].rename(columns={
                "Code": "NCI Term Code"
            }),
            left_on=["Term_norm", "NCI Codelist Code"],
            right_on=["Submission Value", "Codelist Code"],
            how="left"
        )

        codelist_df = codelist_df.drop(
            columns=["Submission Value", "Term_norm", "Codelist Code"],
            errors="ignore"
        )

    # =================================================
    # 9. Decode（只處理 XXTESTCD / XXTEST）
    # =================================================
    if "Decode" not in codelist_df.columns:
        codelist_df["Decode"] = ""

    codelist_df["ID"] = codelist_df["ID"].fillna("").astype(str)

    codelist_df["Decode_Lookup_ID"] = (
        codelist_df["ID"].str.replace("TESTCD", "TEST", regex=False)
    )

    test_name_df = codelist_df[
        codelist_df["ID"].str.endswith("TEST") &
        ~codelist_df["ID"].str.endswith("TESTCD")
    ][[
        "Dataset", "ID", "NCI Term Code", "Term"
    ]].copy()

    test_name_df = test_name_df.rename(columns={
        "ID": "Decode_Lookup_ID",
        "Term": "Decode_from_TEST"
    })

    if not test_name_df.empty:
        codelist_df = codelist_df.merge(
            test_name_df.drop_duplicates(),
            on=["Dataset", "Decode_Lookup_ID", "NCI Term Code"],
            how="left"
        )

    testcd_mask = (
        (codelist_df["ID"].fillna("").astype(str).str.endswith("TESTCD"))
        &
        (codelist_df["NCI Term Code"].fillna("").astype(str).str.strip().ne(""))
    )

    codelist_df.loc[testcd_mask, "Decode"] = (
        codelist_df.loc[testcd_mask, "Decode_from_TEST"].fillna("")
    )

    codelist_df = codelist_df.drop(columns=["Decode_Lookup_ID", "Decode_from_TEST"], errors="ignore")

    # =================================================
    # 10. TSPARMCD Decode（用 CT master）
    # =================================================
    if ct_master_df is not None and not ct_master_df.empty:

        right = ct_master_df.copy()

        for col in ["Submission Value", "Code", "Codelist Name"]:
            if col not in right.columns:
                right[col] = ""

        right["Submission Value"] = (
            right["Submission Value"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )

        right["Code"] = right["Code"].fillna("").astype(str).str.strip()
        right["Codelist Name"] = right["Codelist Name"].fillna("").astype(str).str.strip()

        tsparm_name_df = right[
            right["Codelist Name"] == "Trial Summary Parameter Test Name"
        ][["Code", "Submission Value"]].copy()

        tsparm_name_df = tsparm_name_df.rename(columns={
            "Code": "NCI Term Code",
            "Submission Value": "Decode_TSPARM"
        })

        tsparmcd_mask = (
            (codelist_df["ID"].fillna("").astype(str).str.upper() == "TSPARMCD")
            &
            (codelist_df["NCI Term Code"].fillna("").astype(str).str.strip().ne(""))
        )

        if tsparmcd_mask.any():
            codelist_df = codelist_df.merge(
                tsparm_name_df.drop_duplicates(),
                on="NCI Term Code",
                how="left"
            )

            codelist_df.loc[tsparmcd_mask, "Decode"] = (
                codelist_df.loc[tsparmcd_mask, "Decode_TSPARM"].fillna("")
            )

            codelist_df = codelist_df.drop(columns=["Decode_TSPARM"], errors="ignore")

    # =================================================
    # 11. 排序 / 去重
    # =================================================
    codelist_df = (
        codelist_df
        .drop_duplicates()
        .sort_values(by=["Codelist", "Term", "Dataset", "Variable"], na_position="last")
        .reset_index(drop=True)
    )

    # =================================================
    # Final UI 欄位
    # =================================================
    #Data Type 固定
    codelist_df["Data Type"] = "text"

    # Comment 固定空
    codelist_df["Comment"] = ""

    # Decoded Value
    codelist_df["Decoded Value"] = codelist_df.get("Decode", "")

    # Order（ID分組 + 依Term排序）
    codelist_df = codelist_df.sort_values(["ID", "Term"], kind="stable")

    codelist_df["Order"] = (
        codelist_df
        .groupby("ID")
        .cumcount()
        + 1
    )

    #  清理
    for col in codelist_df.columns:
        codelist_df[col] = codelist_df[col].fillna("").astype(str)


    return codelist_df
    # End=========================================================



def build_dictionaries_sheet(meddra_version="", cm_dictionary="WHO ATC/DDD", cm_version=""):
    return pd.DataFrame([
        {
            "ID": "AEDICT_F",
            "Name": "Adverse Event Dictionary",
            "Data Type": "text",
            "Dictionary": "MEDDRA",
            "Version": meddra_version
        },
        {
            "ID": "CMDICT_F",
            "Name": "Concomitant Meds Dictionary",
            "Data Type": "text",
            "Dictionary": cm_dictionary,
            "Version": cm_version
        },
        {
            "ID": "ISO3166",
            "Name": "Country Codes (ISO 3166)",
            "Data Type": "text",
            "Dictionary": "ISO 3166",
            "Version": ""
        }
    ])
    # End=========================================================





# =================================================================================================================
# 系統流程設定
# =================================================================================================================
# 判斷 Step 1 結果能不能重用（避免每次重跑）
def make_step1_cache_key(file_bytes):
    md5 = hashlib.md5(file_bytes).hexdigest()
    return md5
    # End=========================================================


# =================================================================================================================
# 從上傳的txt中抓CT版本
# =================================================================================================================
def get_available_ct_versions():
    base_path = "config/sdtm_ct"

    files = [
        f for f in os.listdir(base_path)
        if f.startswith("SDTM Terminology") and f.endswith(".txt")
    ]

    version_map = {}

    for f in files:
        # 抓 yyyy-mm-dd
        m = re.search(r"(\d{4}-\d{2}-\d{2})", f)
        if m:
            version = m.group(1)
            version_map[version] = f

    # 排序（新 → 舊）
    versions = sorted(version_map.keys(), reverse=True)

    return versions, version_map
    # End=========================================================




def build_variable_mapping_table(detail_df, variables_spec_df):
    """
    Variable Mapping Table
    目標：
      1. 保留所有 SDTM SPEC variables（以 variables_spec_df 為主）
      2. 補上 detail_df 的 raw mapping
      3. 保留 CRF Dataset / CRF Variable / CRF Data Type
    """
    
    final_cols = [
        "CRF Dataset", "CRF Variable", "CRF Data Type",
        "Order", "Dataset", "Variable", "Label", "Data Type",
        "Codelist", "Origin", "Source"
    ]

    # -------------------------------------------------
    # helper
    # -------------------------------------------------
    def clean_str(x):
        return "" if pd.isna(x) else str(x).strip()

    # -------------------------------------------------
    # 0. 保底：沒有 variables_spec_df 直接回空
    # -------------------------------------------------
    if variables_spec_df is None or variables_spec_df.empty:
        return pd.DataFrame(columns=final_cols)

    # -------------------------------------------------
    # 1. 先整理 variables_spec_df
    #    這張是主體：所有 SDTM SPEC 變數都要留
    # -------------------------------------------------
    v = variables_spec_df.copy()
    v.columns = [str(c).strip() for c in v.columns]

    for c in [
        "Order", "Dataset", "Variable", "Label", "Data Type",
        "Codelist", "Origin", "Source"
    ]:
        if c not in v.columns:
            v[c] = ""

    v["Dataset"] = v["Dataset"].fillna("").astype(str).str.strip().str.upper()
    v["Variable"] = v["Variable"].fillna("").astype(str).str.strip().str.upper()

    # 主 metadata（每個 Dataset+Variable 保留一筆）
    spec_meta = v[[
        "Order", "Dataset", "Variable", "Label", "Data Type",
        "Codelist", "Origin", "Source"
    ]].drop_duplicates(subset=["Dataset", "Variable"], keep="first").reset_index(drop=True)

    # -------------------------------------------------
    # 2. 整理 detail_df（raw source mapping）
    # -------------------------------------------------
    if detail_df is None or detail_df.empty:
        detail_map = pd.DataFrame(columns=[
            "CRF Dataset", "CRF Variable", "CRF Data Type",
            "Dataset", "Variable"
        ])
    else:
        d = detail_df.copy()
        d.columns = [str(c).strip() for c in d.columns]

        for c in ["CRF Dataset", "CRF Variable", "SDTM Domain", "SDTM Variable"]:
            if c not in d.columns:
                d[c] = ""

        # CRF Data Type 保底
        if "CRF Data Type" not in d.columns:
            d["CRF Data Type"] = ""

        d["CRF Dataset"] = d["CRF Dataset"].fillna("").astype(str).str.strip()
        d["CRF Variable"] = d["CRF Variable"].fillna("").astype(str).str.strip()
        d["CRF Data Type"] = d["CRF Data Type"].fillna("").astype(str).str.strip()

        d["SDTM Domain"] = d["SDTM Domain"].fillna("").astype(str).str.strip().str.upper()
        d["SDTM Variable"] = d["SDTM Variable"].fillna("").astype(str).str.strip().str.upper()

        detail_map = d.rename(columns={
            "SDTM Domain": "Dataset",
            "SDTM Variable": "Variable"
        })[[
            "CRF Dataset", "CRF Variable", "CRF Data Type",
            "Dataset", "Variable"
        ]].drop_duplicates().reset_index(drop=True)

    # -------------------------------------------------
    # 3. 用 variables_spec_df 當主體，left merge detail source
    #    → 所有 spec variable 都保留
    # -------------------------------------------------
    out = spec_meta.merge(
        detail_map,
        how="left",
        on=["Dataset", "Variable"]
    )

    # -------------------------------------------------
    # 4. 保底欄位
    # -------------------------------------------------
    for c in final_cols:
        if c not in out.columns:
            out[c] = ""

    for c in ["CRF Dataset", "CRF Variable", "CRF Data Type"]:
        out[c] = out[c].fillna("").astype(str).str.strip()

    # -------------------------------------------------
    # 5. 排序
    # -------------------------------------------------
    out["Order"] = pd.to_numeric(out["Order"], errors="coerce")

    out = out[final_cols].copy()
    out = out.sort_values(
        by=["Dataset", "Order", "Variable", "CRF Dataset", "CRF Variable"],
        na_position="last"
    ).reset_index(drop=True)

    return out
    # End=========================================================



def build_value_mapping_table(
    detail_df,
    variables_spec_df=None,
    ct_mapping_df=None,
    matched_ct_df=None,
    codelist_df=None
):
    """
    Value Mapping:
    只處理 value-level mapping，不處理 ASSIGN
    """

    import pandas as pd

    def clean_str(x):
        return "" if pd.isna(x) else str(x).strip()

    if detail_df is None or detail_df.empty:
        return pd.DataFrame(columns=[
            "CRF Dataset", "CRF Variable",
            "Dataset", "Variable", "Codelist", "CT Code",
            "CRF Option Value", "Original Value", "Original Value Normalized",
            "CT Term", "NCI Term Code", "Decoded Value"
        ])

    d = detail_df.copy()
    d.columns = [str(c).strip() for c in d.columns]

    for c in ["CRF Dataset", "CRF Variable", "SDTM Domain", "SDTM Variable", "Assign Value"]:
        if c not in d.columns:
            d[c] = ""

    d["CRF Dataset"] = d["CRF Dataset"].fillna("").astype(str).str.strip()
    d["CRF Variable"] = d["CRF Variable"].fillna("").astype(str).str.strip()
    d["SDTM Domain"] = d["SDTM Domain"].fillna("").astype(str).str.strip().str.upper()
    d["SDTM Variable"] = d["SDTM Variable"].fillna("").astype(str).str.strip().str.upper()
    d["Assign Value"] = d["Assign Value"].fillna("").astype(str).str.strip()

    # ✅ 只保留非 ASSIGN（value mapping 只看 collected）
    base_df = d[
        d["Assign Value"].fillna("").astype(str).str.strip() == ""
    ].rename(columns={
        "SDTM Domain": "Dataset",
        "SDTM Variable": "Variable"
    })[
        ["CRF Dataset", "CRF Variable", "Dataset", "Variable"]
    ].drop_duplicates().reset_index(drop=True)

    # 補 variables metadata
    if variables_spec_df is not None and not variables_spec_df.empty:
        v = variables_spec_df.copy()
        v.columns = [str(c).strip() for c in v.columns]

        for c in ["Dataset", "Variable", "Codelist", "CT Code"]:
            if c not in v.columns:
                v[c] = ""

        v["Dataset"] = v["Dataset"].fillna("").astype(str).str.strip().str.upper()
        v["Variable"] = v["Variable"].fillna("").astype(str).str.strip().str.upper()
        v["Codelist"] = v["Codelist"].fillna("").astype(str).str.strip().str.upper()
        v["CT Code"] = v["CT Code"].fillna("").astype(str).str.strip().str.upper()

        meta = v[["Dataset", "Variable", "Codelist", "CT Code"]].drop_duplicates(
            subset=["Dataset", "Variable"],
            keep="first"
        )

        base_df = base_df.merge(meta, how="left", on=["Dataset", "Variable"])
    else:
        base_df["Codelist"] = ""
        base_df["CT Code"] = ""

    # ct_mapping_df 補 raw values
    base_df["CRF Option Value"] = ""
    base_df["Original Value"] = ""

    if ct_mapping_df is not None and not ct_mapping_df.empty:
        ct = ct_mapping_df.copy()
        ct.columns = [str(c).strip() for c in ct.columns]

        rename_map = {}
        for c in ct.columns:
            cu = c.upper().strip()
            if cu == "SDTM DOMAIN":
                rename_map[c] = "Dataset"
            elif cu == "SDTM VARIABLE":
                rename_map[c] = "Variable"

        ct = ct.rename(columns=rename_map)

        for c in [
            "CRF Dataset", "CRF Variable", "Dataset", "Variable",
            "CRF Option Value", "Original Value"
        ]:
            if c not in ct.columns:
                ct[c] = ""

        ct["CRF Dataset"] = ct["CRF Dataset"].fillna("").astype(str).str.strip()
        ct["CRF Variable"] = ct["CRF Variable"].fillna("").astype(str).str.strip()
        ct["Dataset"] = ct["Dataset"].fillna("").astype(str).str.strip().str.upper()
        ct["Variable"] = ct["Variable"].fillna("").astype(str).str.strip().str.upper()
        ct["CRF Option Value"] = ct["CRF Option Value"].fillna("").astype(str).str.strip()
        ct["Original Value"] = ct["Original Value"].fillna("").astype(str).str.strip()

        raw_rows = []

        for _, b in base_df.iterrows():
            ds = clean_str(b["Dataset"]).upper()
            var = clean_str(b["Variable"]).upper()
            crf_ds = clean_str(b["CRF Dataset"])
            crf_var = clean_str(b["CRF Variable"])

            subset = ct[
                (ct["Dataset"] == ds) &
                (ct["Variable"] == var) &
                (ct["CRF Dataset"] == crf_ds) &
                (ct["CRF Variable"] == crf_var)
            ].copy()

            if subset.empty:
                row = b.to_dict()
                row["CRF Option Value"] = ""
                row["Original Value"] = ""
                raw_rows.append(row)
            else:
                for _, s in subset.iterrows():
                    row = b.to_dict()
                    row["CRF Option Value"] = clean_str(s.get("CRF Option Value", ""))
                    row["Original Value"] = clean_str(s.get("Original Value", ""))
                    raw_rows.append(row)

        base_df = pd.DataFrame(raw_rows).drop_duplicates().reset_index(drop=True)

    # matched_ct_df 補 CT term
    base_df["Original Value Normalized"] = ""
    base_df["CT Term"] = ""

    if matched_ct_df is not None and not matched_ct_df.empty:
        m = matched_ct_df.copy()
        m.columns = [str(c).strip() for c in m.columns]

        rename_map = {}
        for c in m.columns:
            cu = c.upper().strip()
            if cu == "SDTM DOMAIN":
                rename_map[c] = "Dataset"
            elif cu == "SDTM VARIABLE":
                rename_map[c] = "Variable"

        m = m.rename(columns=rename_map)

        for c in [
            "CRF Dataset", "CRF Variable", "Dataset", "Variable",
            "Original Value", "Original Value Normalized", "CT Term"
        ]:
            if c not in m.columns:
                m[c] = ""

        m["CRF Dataset"] = m["CRF Dataset"].fillna("").astype(str).str.strip()
        m["CRF Variable"] = m["CRF Variable"].fillna("").astype(str).str.strip()
        m["Dataset"] = m["Dataset"].fillna("").astype(str).str.strip().str.upper()
        m["Variable"] = m["Variable"].fillna("").astype(str).str.strip().str.upper()
        m["Original Value"] = m["Original Value"].fillna("").astype(str).str.strip()
        m["Original Value Normalized"] = m["Original Value Normalized"].fillna("").astype(str).str.strip()
        m["CT Term"] = m["CT Term"].fillna("").astype(str).str.strip()

        ct_rows = []

        for _, b in base_df.iterrows():
            ds = clean_str(b["Dataset"]).upper()
            var = clean_str(b["Variable"]).upper()
            crf_ds = clean_str(b["CRF Dataset"])
            crf_var = clean_str(b["CRF Variable"])
            orig = clean_str(b["Original Value"])

            subset = m[
                (m["Dataset"] == ds) &
                (m["Variable"] == var) &
                (m["CRF Dataset"] == crf_ds) &
                (m["CRF Variable"] == crf_var)
            ].copy()

            if orig:
                subset = subset[
                    subset["Original Value"].fillna("").astype(str).str.strip() == orig
                ].copy()

            if subset.empty:
                ct_rows.append(b.to_dict())
            else:
                for _, s in subset.iterrows():
                    row = b.to_dict()
                    row["Original Value"] = clean_str(s.get("Original Value", "")) or row.get("Original Value", "")
                    row["Original Value Normalized"] = clean_str(s.get("Original Value Normalized", ""))
                    row["CT Term"] = clean_str(s.get("CT Term", ""))
                    ct_rows.append(row)

        base_df = pd.DataFrame(ct_rows).drop_duplicates().reset_index(drop=True)

    # codelist enrich
    base_df["NCI Term Code"] = ""
    base_df["Decoded Value"] = ""

    if codelist_df is not None and not codelist_df.empty:
        c = codelist_df.copy()
        c.columns = [str(cn).strip() for cn in c.columns]

        for col in ["ID", "Term", "NCI Term Code", "Decoded Value"]:
            if col not in c.columns:
                c[col] = ""

        c["ID"] = c["ID"].fillna("").astype(str).str.strip().str.upper()
        c["Term"] = c["Term"].fillna("").astype(str).str.strip()
        c["NCI Term Code"] = c["NCI Term Code"].fillna("").astype(str).str.strip()
        c["Decoded Value"] = c["Decoded Value"].fillna("").astype(str).str.strip()

        enrich_rows = []

        for _, b in base_df.iterrows():
            variable = clean_str(b["Variable"]).upper()
            term = clean_str(b.get("CT Term", ""))

            if term == "":
                enrich_rows.append(b.to_dict())
                continue

            subset = c[
                (c["ID"] == variable) &
                (c["Term"] == term)
            ].copy()

            if subset.empty:
                enrich_rows.append(b.to_dict())
            else:
                s = subset.iloc[0]
                row = b.to_dict()
                row["NCI Term Code"] = clean_str(s.get("NCI Term Code", ""))
                row["Decoded Value"] = clean_str(s.get("Decoded Value", ""))
                enrich_rows.append(row)

        base_df = pd.DataFrame(enrich_rows).drop_duplicates().reset_index(drop=True)

    out_cols = [
        "CRF Dataset", "CRF Variable",
        "Dataset", "Variable", "Codelist", "CT Code",
        "CRF Option Value", "Original Value", "Original Value Normalized",
        "CT Term", "NCI Term Code", "Decoded Value"
    ]

    for c in out_cols:
        if c not in base_df.columns:
            base_df[c] = ""

    out_df = base_df[out_cols].copy()

    out_df = out_df.sort_values(
        by=["Dataset", "Variable", "CRF Dataset", "CRF Variable", "Original Value", "CT Term"],
        na_position="last"
    ).reset_index(drop=True)

    return out_df
    # End=========================================================








# =========================================================================================================================================================
# 主流程 UI
# =========================================================================================================================================================

st.set_page_config(page_title="Auto SDTM SPEC", layout="wide")

st.markdown("""
<style>
.stButton > button {
    background-color: black !important;
    color: white !important;
    border-radius: 6px;
    border: none;
}

.stButton > button:hover {
    background-color: #333333 !important;
    color: white !important;
}
</style>
""", unsafe_allow_html=True)


st.title("Auto SDTM SPEC")


uploaded_file = st.file_uploader("請上傳 eCRF Schema Excel", type=["xlsx", "xls"])

if uploaded_file is not None:
    file_bytes = uploaded_file.getvalue()

    current_upload_key = f"{uploaded_file.name}_{uploaded_file.size}"
    if st.session_state.get("current_upload_key") != current_upload_key:
        st.session_state["current_upload_key"] = current_upload_key
        st.session_state["run_step2"] = False
        st.session_state["step1_cache_key"] = None
        st.session_state["step1_result"] = None

    try:
        xls = pd.ExcelFile(BytesIO(file_bytes))
        all_sheets = xls.sheet_names
       
        # -------------------------------------------------
        # 前置作業
        # -------------------------------------------------
        # Sponsor_Protocol
        st.markdown("#### Basic Information")

        sponsor, default_protocol_no = extract_protocol_no_from_filename(uploaded_file.name)

        col1, col2 = st.columns(2)

        with col1:
            protocol_no = st.text_input(
                "Protocol No",
                value=default_protocol_no,
                key="protocol_no"
            )

        with col2:
            protocol_title = st.text_input(
                "Protocol Title",
                value="",
                key="protocol_title"
            )

        # Version Control
        st.markdown("#### Version Control")

        r1_c1, r1_c2 = st.columns(2)

        with r1_c1:
            version = st.selectbox(
                "SDTM IG",
                ["Version 3.4", "Version 3.3"],
                key="sdtm_version_selector"
            )

        versions, version_map = get_available_ct_versions()
        with r1_c2:
            sdtm_ct = st.selectbox(
                "SDTM CT",
                options=versions,
                index=0,  # 預設選最新
                key="sdtm_ct"
            )
        
        meddra_version = st.text_input("MedDRA", value="", key="meddra_version")

        r3_c1, r3_c2 = st.columns(2)

        with r3_c1:
            cm_dictionary = st.selectbox(
                "CM 字典",
                ["WHODrug Global B3", "WHO ATC/DDD"],
                key="cm_dictionary"
            )

        with r3_c2:
            cm_version = st.text_input("CM 版本", value="", key="cm_version")

        
        r4_c1, r4_c2, r4_c3 = st.columns(3)

        with r4_c1:
            snomed_version = st.text_input("SNOMED", value="", key="snomed_version")

        with r4_c2:
            unii_version = st.text_input("UNII", value="", key="unii_version")

        with r4_c3:
            medrt_version = st.text_input("MED-RT", value="", key="medrt_version")

        sdtm_ct = normalize_date_text(sdtm_ct)
        snomed_version = normalize_date_text(snomed_version)
        unii_version = normalize_date_text(unii_version)
        medrt_version = normalize_date_text(medrt_version)


        # Load Config
        if (
            not st.session_state.get("config_loaded")
            or st.session_state.get("config_version") != version
        ):

            raw_cfg_df, cfg_path = load_domains_config(version)

            cfg_df = standardize_domains_config(raw_cfg_df)
    
            # 存 config
            st.session_state["config_df"] = cfg_df

            # 建 mapping（CT mapping會用）
            st.session_state["var_to_ctcode"] = dict(
                zip(cfg_df["Variable"], cfg_df["CT Code"])
            )

            # 記錄版本
            st.session_state["config_version"] = version
            st.session_state["config_loaded"] = True
        

        # Load CT Term Mapping Dictionary
        ct_mapping_dict_df = None

        dict_path = "config/CT Term Mapping Dictionary.xlsx"

        if os.path.exists(dict_path):
            try:
                ct_mapping_dict_df = pd.read_excel(dict_path)
                ct_mapping_dict_df = normalize_columns(ct_mapping_dict_df)

                st.session_state["ct_mapping_dict_df"] = ct_mapping_dict_df

                # st.success("CT Mapping Dictionary 載入成功")
            except Exception as e:
                st.warning(f"CT Mapping Dictionary 載入失敗：{e}")
        else:
            st.warning("找不到 CT Mapping Dictionary（目前只使用 seed）")


        
        # -------------------------------------------------
        # Step 1：CRF → SDTM Mapping
        # -------------------------------------------------
        st.markdown("## Step 1｜CRF → SDTM Mapping")

        step1_cache_key = make_step1_cache_key(file_bytes)

        if (
            st.session_state.get("step1_cache_key") == step1_cache_key
            and st.session_state.get("step1_result") is not None
        ):
            result = st.session_state["step1_result"]
        else:
            result = process_uploaded_excel(
                file_bytes=file_bytes,
                all_sheets=all_sheets
            )
            st.session_state["step1_cache_key"] = step1_cache_key
            st.session_state["step1_result"] = result


        # 呼叫SoA
        soa_df = result["soa_list_df"]

        # Visit去重複 (供後續TV使用)
        unique_visit_df = (           
            soa_df
            .loc[
                (soa_df["Visit"].notna()) &
                (soa_df["Visit"].str.strip() != "") &
                (soa_df["CRF Dataset"] != soa_df["Abbreviation"])
            ]
            [["Abbreviation", "Visit", "Visit_order"]]
            .drop_duplicates()
            .sort_values("Visit_order")
            .reset_index(drop=True)
        )
        
        st.session_state["unique_visit_df"] = unique_visit_df
        # st.write(soa_df)

        
        missing_sheets = result["missing_sheets"]
        mapping_df = result["mapping_df"]
        detail_df = result["detail_df"]
        sheet_errors = result["sheet_errors"]
        unparsed_records = result["unparsed_records"]
        ct_mapping_df = result.get("ct_mapping_df", pd.DataFrame())
        ct_mapping_sheet_errors = result.get("ct_mapping_sheet_errors", [])



        # SDTM Varialbe Mapping (Summary by Domain）
        st.markdown("### 📊 SDTM Variable Mapping")
        st.markdown("##### - Summary by Domain")
        
        if mapping_df.empty:
            st.warning("目前沒有從 CRF Sheet 抓到可解析的 SDTM Domain / Variable")
        else:
            summary_df = (
                mapping_df
                .groupby("SDTM Domain")["SDTM Variable"]
                .apply(lambda x: sorted(set(x)))
                .reset_index()
            )

            summary_df["Variable Count"] = summary_df["SDTM Variable"].apply(len)
            summary_df["Variables"] = summary_df["SDTM Variable"].apply(lambda x: "; ".join(x))

            st.dataframe(summary_df[["SDTM Domain", "Variable Count", "Variables"]], use_container_width=True)


        # Detail（CRF → SDTM）
        st.markdown("##### - Detail")
                
        if detail_df.empty:
            st.info("目前沒有可顯示的明細")
        else:          
            sorted_detail_df = detail_df.sort_values(
                by=["SDTM Domain", "SDTM Variable", "CRF Dataset", "CRF Variable"],
                ascending=[True, True, True, True]
            ).reset_index(drop=True)

            st.dataframe(sorted_detail_df, use_container_width=True)

       

        # CT Mapping Result       
        st.markdown("### 🧩 CT Mapping List")

        st.dataframe(ct_mapping_df, use_container_width=True)


        # 確保 dictionary 已載入
        if "ct_mapping_dict_df" in st.session_state:

            mapping_dict_df = st.session_state["ct_mapping_dict_df"]

            matched_ct_df, unmatched_ct_df = build_ct_mapping(
                ct_mapping_df,
                mapping_dict_df
            )

            tab1, tab2, tab3 = st.tabs(["✅ Matched", "❌ Unmatched", "🟡 No CT Code"])

            # -------------------------------------------------
            # ✅ Matched
            # -------------------------------------------------
            with tab1:

                if matched_ct_df.empty:
                    st.info("目前沒有 Matched CT Mapping")
                else:
                    matched_ct_df_filtered = matched_ct_df[
                        matched_ct_df["CT Code"].fillna("").str.strip() != ""
                    ].copy()

                    display_cols = [
                        "SDTM Domain",
                        "SDTM Variable",
                        "CT Code",
                        "Original Value",
                        "Original Value Normalized",
                        "CT Term",
                    ]

                    display_cols = [c for c in display_cols if c in matched_ct_df_filtered.columns]

                    st.dataframe(
                        matched_ct_df_filtered[display_cols],
                        use_container_width=True
                    )


            # -------------------------------------------------
            # ❌ Unmatched
            # -------------------------------------------------
            with tab2:
    
                if unmatched_ct_df.empty:
                    st.success("🎉 所有 CRF Term 都已成功 Mapping")
                else:
                    
                    unmatched_ct_df_filtered = unmatched_ct_df[
                        unmatched_ct_df["CT Code"].fillna("").str.strip() != "NY"
                    ].copy()

                    st.warning("以下 CRF Term 尚未對應 CT Term，建議加入 CT Mapping Dictionary")

                    display_cols = [
                        "SDTM Domain",
                        "SDTM Variable",
                        "CT Code",
                        "Original Value"
                    ]

                    display_cols = [c for c in display_cols if c in unmatched_ct_df_filtered.columns]

                    st.dataframe(
                        unmatched_ct_df_filtered[display_cols],
                        use_container_width=True
                    )

                    # 開發者直接匯出用
                    st.download_button(
                        label="⬇️ 下載 Unmatched CT（通知開發者更新 Mapping Dictionary）",
                        data=unmatched_ct_df_filtered.to_csv(index=False),
                        file_name="ct_mapping_unmatched.csv",
                        mime="text/csv"
                    )
                    
            # -------------------------------------------------
            # 🟡 No CT Code（新增🔥）
            # -------------------------------------------------
            with tab3:

                if ct_mapping_df is None or ct_mapping_df.empty:
                    st.info("沒有 CT Mapping 資料")
                else:

                    no_ct_df = ct_mapping_df[
                        ct_mapping_df["CT Code"].fillna("").str.strip() == ""
                    ].copy()

                    if no_ct_df.empty:
                        st.success("🎉 沒有 CT Code 為空的資料")
                    else:
                        st.warning("以下變數在 Config 沒有定義 CT Code")

                        display_cols = [
                            "SDTM Domain",
                            "SDTM Variable",
                            "Assign Value",
                            "CRF Option Value",
                            "Original Value"
                        ]

                        display_cols = [c for c in display_cols if c in no_ct_df.columns]

                        st.dataframe(
                            no_ct_df[display_cols],
                            use_container_width=True
                        )


        else:
            st.info("尚未載入 CT Mapping Dictionary")



        
        # 錯誤 / Debug
        st.markdown("### ⚠️ Debug / Error 檢查")

        if missing_sheets:
            st.warning(f"SoA 有出現的 Form OID，但 Excel 沒有對應的 Sheets: {missing_sheets}")
        
        if sheet_errors:
            st.warning(f"無法處理的 Sheets (Header偵測失敗): {sorted(set(sheet_errors))}")
        
        if unparsed_records:
            st.markdown("#### 無法解析的 SDTM IG Target")
            st.dataframe(pd.DataFrame(unparsed_records), use_container_width=True)
        
        if ct_mapping_sheet_errors:
            st.warning(f"CT Mapping 無法處理的 Sheets：{sorted(set(ct_mapping_sheet_errors))}")


        
        # -------------------------------------------------
        # Step 2 開關：執行 / 重新整理
        # -------------------------------------------------
        def trigger_step2():
            st.session_state["run_step2"] = True

        st.button(
            "▶ 執行 / 重新整理 Step 2：SPEC Generator",
            type="primary",
            on_click=trigger_step2
        )

        # -------------------------------------------------
        # Step 2：SPEC Generator
        # -------------------------------------------------
        if st.session_state.get("run_step2", False):
            st.markdown("## Step 2｜SPEC Generator")

            # 2.1 Define
            st.markdown("### 2.1 Define")
            define_df = build_define_sheet(
                protocol_no=protocol_no,
                protocol_title=protocol_title,
                sdtm_version=version
            )
            st.dataframe(define_df, use_container_width=True)


            # 先產出5T
            td_dict = build_trial_design_sheets(
                protocol_no=protocol_no,
                protocol_title=protocol_title,
                sdtm_version=version,
                sdtm_ct=sdtm_ct,
                snomed_version=snomed_version,
                medrt_version=medrt_version,
                unii_version=unii_version,
                unique_visit_df=st.session_state.get("unique_visit_df", pd.DataFrame()),
                ct_master_df=st.session_state.get("ct_master_df")
            )
            
            ta_df = td_dict.get("TA", pd.DataFrame())
            te_df = td_dict.get("TE", pd.DataFrame())
            ti_df = td_dict.get("TI", pd.DataFrame())
            ts_df = td_dict.get("TS", pd.DataFrame())
            tv_df = td_dict.get("TV", pd.DataFrame())
            

            # 先產出2.3的Variable List
            variables_spec_df = build_variables_sheet(
                detail_df=detail_df,
                config_df=st.session_state["config_df"],
                td_dict=td_dict
            )       

            variables_view_df = variables_spec_df.drop(columns=["CT Code"])

            
            # 2.2 Datasets
            st.markdown("### 2.2 Datasets")
            
            datasets_df = build_datasets_from_variables(
                variables_spec_df,
                st.session_state["config_df"],
                version
            )
            
            st.dataframe(datasets_df, use_container_width=True)
            
            # 2.3 Variables
            st.markdown("### 2.3 Variables")           
            st.dataframe(variables_view_df, use_container_width=True)
            
            
            # 2.4 Codelists
            st.markdown("### 2.4 Codelists")

            # Load CT Master
            try:
                ct_df, info = load_ct_master(
                    st.session_state.get("sdtm_ct", "")
                )
                
                st.session_state["ct_master_df"] = ct_df
                st.session_state["ct_master_info"] = info

                st.success(f"✅ 使用 SDTM CT：{info['resolved_version']}")
                
                with st.expander("Preview CT Master"):
                    st.dataframe(ct_df.head(20), use_container_width=True)
                
            except Exception as e:
                st.error("❌ SDTM CT 載入失敗")
                st.write(str(e))

            sdtm_ct = info.get("resolved_version", "")
            codelist_df = build_codelist_sheet(
                variables_spec_df=variables_spec_df,
                ct_master_df=st.session_state.get("ct_master_df"),
                matched_ct_df=matched_ct_df,
                ct_mapping_df=ct_mapping_df,
                ts_df=ts_df,
                sdtm_ct=sdtm_ct
            )
            
            display_cols = [
                "ID",
                "Name",
                "NCI Codelist Code",
                "Data Type",
                "Terminology",
                "Comment",
                "Order",
                "Term",
                "NCI Term Code",
                "Decoded Value"
            ]

            # 防欄位不存在（避免Error）
            display_cols = [c for c in display_cols if c in codelist_df.columns]
            
            codelists_export = codelist_df.copy()
            codelists_export = codelists_export[
                [c for c in display_cols if c in codelists_export.columns]
            ]

            st.dataframe(codelists_export, use_container_width=True)


            # 2.5 Codelists
            st.markdown("### 2.5 Dictionaries")
            dictionaries_df = st.data_editor(
                build_dictionaries_sheet(
                    meddra_version=meddra_version,
                    cm_dictionary=cm_dictionary,
                    cm_version=cm_version
                ),
                num_rows="dynamic",
                use_container_width=True,
                key="dictionaries_editor"
            )
            
            # 2.6 Trial Design
            st.markdown("### 2.6 Trial Design (5T)")

            with st.expander("TA / TE / TI / TS / TV 基本欄位骨架", expanded=False):
                st.markdown("#### TA")
                st.dataframe(ta_df, use_container_width=True)

                st.markdown("#### TE")
                st.dataframe(te_df, use_container_width=True)
                
                st.markdown("#### TI")
                st.dataframe(ti_df, use_container_width=True)
                
                st.markdown("#### TS")
                st.dataframe(ts_df, use_container_width=True, height=500)
                
                st.markdown("#### TV")
                st.dataframe(tv_df, use_container_width=True)


            # ==========================================================
            # Excel Sheets
            # ==========================================================
            export_sheets = {
                "Define": define_df,
                "Datasets": datasets_df,
                "Variables": variables_view_df,
                "Codelist": codelists_export,
                "Dictionaries": dictionaries_df,
                "TA": ta_df,
                "TE": te_df,
                "TI": ti_df,
                "TS": ts_df,
                "TV": tv_df
            }

            excel_bytes = Export_excel(export_sheets)

            # 檔名
            today_str = datetime.now().strftime("%Y%m%d")
            file_name = f"{sponsor}_{protocol_no}_Mock SDTM_SPEC_{today_str}.xlsx"

            # Download Button
            st.download_button(
                label="下載 SDTM SPEC Excel",
                data=excel_bytes,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    

            variable_mapping_df = build_variable_mapping_table(
                detail_df=detail_df,
                variables_spec_df=variables_spec_df
            )

            st.markdown("### 🧩 Variable Mapping Table")

            var_display_cols = [
                "CRF Dataset", "CRF Variable", "CRF Data Type",
                "Order", "Dataset", "Variable", "Label", "Data Type",
                "Codelist", "Origin", "Source"
            ]

            var_display_cols = [c for c in var_display_cols if c in variable_mapping_df.columns]

            st.dataframe(
                variable_mapping_df[var_display_cols],
                use_container_width=True,
                height=450
            )



            value_mapping_df = build_value_mapping_table(
                detail_df=detail_df,
                variables_spec_df=variables_spec_df,
                ct_mapping_df=ct_mapping_df,
                matched_ct_df=matched_ct_df,
                codelist_df=codelist_df
            )

            st.markdown("### 🧩 Value Mapping Table")

            val_display_cols = [
                "CRF Dataset", "CRF Variable",
                "Dataset", "Variable", "Codelist", "CT Code",
                "CRF Option Value", "Original Value", "Original Value Normalized",
                "CT Term", "NCI Term Code", "Decoded Value"
            ]

            val_display_cols = [c for c in val_display_cols if c in value_mapping_df.columns]

            st.dataframe(
                value_mapping_df[val_display_cols],
                use_container_width=True,
                height=450
            )

        
    except Exception as e:
        st.error(f"讀取檔案時發生錯誤：{e}")
        st.text(traceback.format_exc())
